import os
import json
import base64
from io import BytesIO
from datetime import datetime, timezone
from urllib.parse import quote, unquote

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from openai import OpenAI
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SMART_REVIEW_IMPORT_ERROR = None

try:
    from components.code_review_ui import render_code_review_results
except Exception as e:
    SMART_REVIEW_IMPORT_ERROR = f"components import failed: {e}"

    def render_code_review_results(review_result):
        st.error("Code review UI could not load.")
        st.write(SMART_REVIEW_IMPORT_ERROR)
        st.json(review_result)

try:
    from services.code_review_service import run_smart_code_review
except Exception as e:
    existing = SMART_REVIEW_IMPORT_ERROR or ""
    SMART_REVIEW_IMPORT_ERROR = f"{existing}\nservices import failed: {e}".strip()

    def run_smart_code_review(*args, **kwargs):
        return {
            "success": False,
            "error": SMART_REVIEW_IMPORT_ERROR,
            "summary": "Smart code review service is unavailable.",
            "issues": [],
            "recommendations": [],
        }

# ------------------------------
# Load environment
# ------------------------------
load_dotenv()

# ------------------------------
# Page config
# ------------------------------
st.set_page_config(
    page_title="Kaldi One",
    page_icon="🧪",
    layout="wide",
)

# ------------------------------
# Clients / secrets
# ------------------------------
def get_secret_or_env(name: str, default=None):
    try:
        return st.secrets[name]
    except Exception:
        return os.getenv(name, default)


OPENAI_API_KEY = get_secret_or_env("OPENAI_API_KEY")
SUPABASE_URL = get_secret_or_env("SUPABASE_URL")
SUPABASE_KEY = get_secret_or_env("SUPABASE_KEY")

if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY not found in Streamlit secrets or .env.")
    st.stop()

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("SUPABASE_URL or SUPABASE_KEY not found in Streamlit secrets or .env.")
    st.stop()

client = OpenAI(api_key=OPENAI_API_KEY)


@st.cache_resource
def get_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


supabase = get_supabase()


# ------------------------------
# Workspace constants
# ------------------------------
QA_WORKSPACE = "QA Workspace"
BA_WORKSPACE = "BA Workspace"
DEV_WORKSPACE = "Dev Workspace"
FLOW_WORKSPACE = "Flow to Requirement"

WORKSPACE_OUTPUT_KEYS = {
    QA_WORKSPACE: "qa_output_state",
    BA_WORKSPACE: "ba_output_state",
    DEV_WORKSPACE: "dev_output_state",
    FLOW_WORKSPACE: "flow_output_state",
}


# ------------------------------
# Browser cookie helpers
# ------------------------------
def set_browser_cookie(name: str, value: str):
    safe_value = quote(value, safe="")
    components.html(
        f"""
        <script>
        document.cookie = "{name}={safe_value}; path=/; SameSite=Lax";
        </script>
        """,
        height=0,
        width=0,
    )



def delete_browser_cookie(name: str):
    components.html(
        f"""
        <script>
        document.cookie = "{name}=; path=/; expires=Thu, 01 Jan 1970 00:00:00 GMT; SameSite=Lax";
        </script>
        """,
        height=0,
        width=0,
    )


# ------------------------------
# Clean UI styling
# ------------------------------
st.markdown("""
<style>
.block-container {
    max-width: 1100px;
    margin: 0 auto;
    padding-top: 3.2rem;
    padding-left: 1rem;
    padding-right: 1rem;
}

.hero-title {
    font-size: 2.3rem;
    font-weight: 700;
    color: #1f2937;
    margin-top: 0.1rem;
    margin-bottom: 0.2rem;
    line-height: 1.2;
}

.hero-subtitle {
    color: #6b7280;
    font-size: 1rem;
    margin-bottom: 1rem;
}

.clean-card {
    border: 1px solid #e5e7eb;
    border-radius: 18px;
    background: #ffffff;
    padding: 1.2rem 1.2rem 1rem 1.2rem;
    margin-bottom: 1rem;
    box-shadow: 0 4px 18px rgba(0,0,0,0.03);
}

.result-card {
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    background: #ffffff;
    padding: 1rem;
    margin-top: 1rem;
}

.small-muted {
    color: #6b7280;
    font-size: 0.86rem;
}

.mermaid-wrap {
    border: 1px solid #e5e7eb;
    border-radius: 14px;
    padding: 0.75rem;
    background: #fafafa;
    margin-bottom: 1rem;
}

section[data-testid="stSidebar"] .stButton > button {
    border-radius: 10px;
}

section[data-testid="stSidebar"] .stSelectbox,
section[data-testid="stSidebar"] .stTextInput {
    margin-bottom: 0.35rem;
}

.top-header-row {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 16px;
    margin-top: 0.6rem;
    margin-bottom: 0.5rem;
}

.top-header-left {
    flex: 1;
    min-width: 0;
}

.top-header-right {
    display: flex;
    justify-content: flex-end;
    align-items: center;
    gap: 10px;
    flex-wrap: nowrap;
    margin-top: 0.15rem;
}

.top-link-btn {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    text-decoration: none !important;
    border-radius: 10px;
    padding: 8px 14px;
    font-size: 0.85rem;
    font-weight: 600;
    line-height: 1;
    white-space: nowrap;
    border: 1px solid #cbd5e1;
    transition: all 0.15s ease;
}

.top-link-btn.pro {
    background: #2563eb;
    color: #ffffff !important;
    border-color: #2563eb;
}

.top-link-btn.pro:hover {
    background: #1d4ed8;
    border-color: #1d4ed8;
}

.top-link-btn.support {
    background: #f8fafc;
    color: #334155 !important;
}

.top-link-btn.support:hover {
    background: #f1f5f9;
    border-color: #94a3b8;
}

.top-link-btn.logout {
    background: #ffffff;
    color: #475569 !important;
}

.top-link-btn.logout:hover {
    background: #f8fafc;
    border-color: #94a3b8;
}

div.stButton > button[kind="primary"] {
    background: #2563eb;
    border-color: #2563eb;
}

div.stButton > button[kind="primary"]:hover {
    background: #1d4ed8;
    border-color: #1d4ed8;
}
</style>
""", unsafe_allow_html=True)


# ------------------------------
# Utility helpers
# ------------------------------
def safe_filename(text):
    text = "" if text is None else str(text)
    cleaned = "".join(c if c.isalnum() or c in (" ", "_", "-") else "" for c in text)
    cleaned = cleaned.strip().replace(" ", "_").lower()
    return cleaned if cleaned else "ai_output"



def clean_text_for_storage(text):
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)

    text = text.replace("\x00", "")

    cleaned = []
    for ch in text:
        if ch in ("\n", "\r", "\t") or ord(ch) >= 32:
            cleaned.append(ch)

    return "".join(cleaned)



def trim_text_for_prompt(text: str, max_chars: int = 15000) -> str:
    text = clean_text_for_storage(text)
    if not text:
        return ""
    text = text.strip()
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n\n[Content truncated because uploaded file was too large.]"



def now_iso():
    return datetime.now(timezone.utc).isoformat()



def parse_json_response(content):
    content = content.strip()
    if content.startswith("```"):
        content = content.replace("```json", "").replace("```", "").strip()
    return json.loads(content)



def encode_uploaded_image(uploaded_file):
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    encoded = base64.b64encode(file_bytes).decode("utf-8")
    mime_type = uploaded_file.type
    return f"data:{mime_type};base64,{encoded}"



def read_uploaded_text_file(uploaded_file):
    if uploaded_file is None:
        return ""
    try:
        file_bytes = uploaded_file.read()
        uploaded_file.seek(0)
        text = file_bytes.decode("utf-8", errors="ignore")
        return clean_text_for_storage(text)
    except Exception:
        return ""



def convert_df_to_csv(df):
    return df.to_csv(index=False).encode("utf-8")



def convert_df_to_excel(df, sheet_name="AI_Output"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        header_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                cell.alignment = wrap_alignment
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 60)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        worksheet.freeze_panes = "A2"

    return output.getvalue()


ALL_UPLOAD_FILE_TYPES = [
    "pdf", "doc", "docx", "xls", "xlsx", "csv", "txt", "md",
    "png", "jpg", "jpeg"
]

DEV_UPLOAD_FILE_TYPES = [
    "py", "js", "ts", "java", "cs", "sql", "txt", "json", "md",
    "pdf", "doc", "docx", "xls", "xlsx", "csv", "png", "jpg", "jpeg"
]


# ------------------------------
# Session state init
# ------------------------------
def build_empty_output_state():
    return {
        "generated_type": None,
        "generated_text": "",
        "generated_df": None,
        "generated_title": "",
        "generated_base_name": "",
        "generated_sheet_name": "",
        "generated_mermaid_code": "",
        "original_output_text": "",
        "editable_output_text": "",
        "original_output_df": None,
        "editable_output_df": None,
        "smart_review_result": None,
        "uploaded_name": "",
    }



def init_session_state():
    defaults = {
        "user": None,
        "access_token": None,
        "refresh_token": None,
        "selected_project_id": None,
        "auth_checked": False,
        "active_workspace": QA_WORKSPACE,
        WORKSPACE_OUTPUT_KEYS[QA_WORKSPACE]: build_empty_output_state(),
        WORKSPACE_OUTPUT_KEYS[BA_WORKSPACE]: build_empty_output_state(),
        WORKSPACE_OUTPUT_KEYS[DEV_WORKSPACE]: build_empty_output_state(),
        WORKSPACE_OUTPUT_KEYS[FLOW_WORKSPACE]: build_empty_output_state(),
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value.copy() if isinstance(value, dict) else value


init_session_state()



def get_workspace_state(workspace_name: str) -> dict:
    key = WORKSPACE_OUTPUT_KEYS[workspace_name]
    if key not in st.session_state:
        st.session_state[key] = build_empty_output_state()
    return st.session_state[key]



def reset_workspace_state(workspace_name: str):
    st.session_state[WORKSPACE_OUTPUT_KEYS[workspace_name]] = build_empty_output_state()



def maybe_reset_workspace_for_new_upload(workspace_name: str, uploaded_file):
    if uploaded_file is None:
        return
    state = get_workspace_state(workspace_name)
    new_name = uploaded_file.name
    old_name = state.get("uploaded_name", "")
    if old_name and old_name != new_name:
        reset_workspace_state(workspace_name)
        state = get_workspace_state(workspace_name)
    state["uploaded_name"] = new_name


# ------------------------------
# Copy / render helpers
# ------------------------------
def render_copy_button(text, button_label="Copy Output"):
    safe_text = json.dumps(text)
    components.html(
        f"""
        <button
            onclick='navigator.clipboard.writeText({safe_text})'
            style="
                background:#2563eb;
                color:white;
                border:none;
                border-radius:10px;
                padding:10px 16px;
                font-weight:600;
                cursor:pointer;
                width:100%;
            "
        >
            {button_label}
        </button>
        """,
        height=52,
    )



def render_mermaid_diagram_with_exports(mermaid_code: str, key_suffix: str, diagram_title: str, height: int = 620):
    if not mermaid_code.strip():
        return

    unique_key = f"mermaid_{safe_filename(key_suffix)}_{int(datetime.now().timestamp() * 1000)}"
    safe_title_js = json.dumps(diagram_title or "diagram")
    safe_filename_js = json.dumps(f"{safe_filename(diagram_title or key_suffix)}")
    safe_mermaid_js = json.dumps(mermaid_code)

    html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8" />
    <style>
        body {{
            margin: 0;
            padding: 12px;
            background: white;
            font-family: Arial, sans-serif;
        }}
        .panel {{
            border: 1px solid #e5e7eb;
            border-radius: 14px;
            background: #ffffff;
            overflow: hidden;
        }}
        .toolbar {{
            display: flex;
            gap: 8px;
            padding: 12px;
            border-bottom: 1px solid #e5e7eb;
            background: #f8fafc;
            flex-wrap: wrap;
        }}
        .toolbar button {{
            border: 1px solid #cbd5e1;
            background: white;
            border-radius: 8px;
            padding: 8px 12px;
            cursor: pointer;
            font-weight: 600;
        }}
        .toolbar button.primary {{
            background: #2563eb;
            border-color: #2563eb;
            color: white;
        }}
        .canvas {{
            padding: 16px;
            background: #fafafa;
            overflow: auto;
            min-height: 420px;
        }}
        .canvas svg {{
            max-width: 100%;
            height: auto;
            background: white;
        }}
        .status {{
            padding: 0 16px 12px 16px;
            color: #64748b;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="panel">
        <div class="toolbar">
            <button class="primary" onclick="downloadPdf_{unique_key}()">Download PDF</button>
            <button onclick="downloadSvg_{unique_key}()">Download SVG</button>
            <button onclick="printDiagram_{unique_key}()">Print</button>
        </div>
        <div class="canvas">
            <div id="diagram_{unique_key}"></div>
        </div>
        <div class="status" id="status_{unique_key}">Rendering diagram...</div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/jspdf@2.5.1/dist/jspdf.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/svg2pdf.js@2.5.0/dist/svg2pdf.umd.min.js"></script>
    <script>
        const mermaidCode_{unique_key} = {safe_mermaid_js};
        const fileBase_{unique_key} = {safe_filename_js};
        const title_{unique_key} = {safe_title_js};
        const statusEl_{unique_key} = document.getElementById("status_{unique_key}");
        const container_{unique_key} = document.getElementById("diagram_{unique_key}");

        mermaid.initialize({{ startOnLoad: false, securityLevel: 'loose' }});

        async function renderDiagram_{unique_key}() {{
            try {{
                const renderResult = await mermaid.render('svg_{unique_key}', mermaidCode_{unique_key});
                container_{unique_key}.innerHTML = renderResult.svg;
                statusEl_{unique_key}.textContent = 'Ready. You can download PDF or SVG.';
            }} catch (err) {{
                statusEl_{unique_key}.textContent = 'Failed to render diagram: ' + err;
            }}
        }}

        function getSvg_{unique_key}() {{
            return container_{unique_key}.querySelector('svg');
        }}

        window.downloadSvg_{unique_key} = function() {{
            const svg = getSvg_{unique_key}();
            if (!svg) {{
                alert('Diagram is not ready yet.');
                return;
            }}
            const blob = new Blob([svg.outerHTML], {{ type: 'image/svg+xml;charset=utf-8' }});
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = fileBase_{unique_key} + '.svg';
            document.body.appendChild(a);
            a.click();
            a.remove();
            URL.revokeObjectURL(url);
        }};

        window.downloadPdf_{unique_key} = async function() {{
            const svg = getSvg_{unique_key}();
            if (!svg) {{
                alert('Diagram is not ready yet.');
                return;
            }}

            try {{
                const {{ jsPDF }} = window.jspdf;
                const svgBox = svg.getBBox();
                const margin = 20;
                const pdfWidth = Math.max(svgBox.width + margin * 2, 595.28);
                const pdfHeight = Math.max(svgBox.height + margin * 2 + 30, 841.89);
                const pdf = new jsPDF({{
                    orientation: pdfWidth > pdfHeight ? 'landscape' : 'portrait',
                    unit: 'pt',
                    format: [pdfWidth, pdfHeight],
                }});

                pdf.setFontSize(14);
                pdf.text(title_{unique_key}, margin, 18);
                await pdf.svg(svg, {{ x: margin, y: 30, width: svgBox.width, height: svgBox.height }});
                pdf.save(fileBase_{unique_key} + '.pdf');
            }} catch (err) {{
                alert('PDF download failed: ' + err);
            }}
        }};

        window.printDiagram_{unique_key} = function() {{
            const svg = getSvg_{unique_key}();
            if (!svg) {{
                alert('Diagram is not ready yet.');
                return;
            }}
            const printWindow = window.open('', '_blank');
            printWindow.document.write(`
                <html>
                    <head>
                        <title>${{title_{unique_key}}}</title>
                        <style>
                            body {{ font-family: Arial, sans-serif; padding: 20px; }}
                            svg {{ max-width: 100%; height: auto; }}
                        </style>
                    </head>
                    <body>
                        <h2>${{title_{unique_key}}}</h2>
                        ${{svg.outerHTML}}
                    </body>
                </html>
            `);
            printWindow.document.close();
            printWindow.focus();
            printWindow.print();
        }};

        renderDiagram_{unique_key}();
    </script>
</body>
</html>
"""

    components.html(html_content, height=height, scrolling=True)



def show_mermaid_download_buttons(mermaid_code: str, base_name: str, diagram_title: str):
    if not mermaid_code.strip():
        return

    d1 = st.columns(1)[0]
    with d1:
        st.download_button(
            label="Download Mermaid Code (.mmd)",
            data=mermaid_code,
            file_name=f"{base_name}_diagram.mmd",
            mime="text/plain",
            use_container_width=True,
            key=f"download_mermaid_code_{base_name}",
        )

    st.caption("Use the diagram toolbar above to download PDF or SVG directly.")


# ------------------------------
# Auth helpers
# ------------------------------
def sign_up_user(email: str, password: str):
    return supabase.auth.sign_up({"email": email, "password": password})



def sign_in_user(email: str, password: str):
    return supabase.auth.sign_in_with_password({"email": email, "password": password})


PASSWORD_RESET_REDIRECT = get_secret_or_env("PASSWORD_RESET_REDIRECT", "http://localhost:8501")



def send_password_reset_email(email: str):
    return supabase.auth.reset_password_for_email(
        email,
        {"redirect_to": PASSWORD_RESET_REDIRECT},
    )



def update_logged_in_user_password(new_password: str):
    return supabase.auth.update_user({"password": new_password})



def auth_error_text(exc: Exception) -> str:
    text = str(exc)
    lower = text.lower()

    if "invalid login credentials" in lower:
        return "Unable to sign in. New user? Please use Sign Up. Already have an account? Use Forgot Password if needed."
    if "email not confirmed" in lower:
        return "Your account is created, but your email is not confirmed yet. Please check your inbox and confirm your email."
    if "user already registered" in lower:
        return "This account already exists. Please use Sign In or Forgot Password."
    if "invalid api key" in lower:
        return "Invalid Supabase configuration. Please check your Supabase URL and key in secrets.toml."
    return text



def handle_login_success(auth_response):
    if getattr(auth_response, "user", None):
        st.session_state.user = auth_response.user

    if getattr(auth_response, "session", None):
        st.session_state.access_token = auth_response.session.access_token
        st.session_state.refresh_token = auth_response.session.refresh_token
        set_browser_cookie("sb_access_token", auth_response.session.access_token)
        set_browser_cookie("sb_refresh_token", auth_response.session.refresh_token)

    upsert_profile(st.session_state.user)
    ensure_default_project(st.session_state.user.id)
    st.rerun()



def sign_out_user():
    try:
        supabase.auth.sign_out()
    except Exception:
        pass

    delete_browser_cookie("sb_access_token")
    delete_browser_cookie("sb_refresh_token")

    for key in list(st.session_state.keys()):
        del st.session_state[key]

    init_session_state()



def load_user_from_existing_session():
    try:
        access_token = st.context.cookies.get("sb_access_token")
        refresh_token = st.context.cookies.get("sb_refresh_token")

        if access_token:
            access_token = unquote(access_token)
        if refresh_token:
            refresh_token = unquote(refresh_token)

        if access_token and refresh_token:
            session_response = supabase.auth.set_session(access_token, refresh_token)

            if session_response and getattr(session_response, "user", None):
                st.session_state.user = session_response.user

            if session_response and getattr(session_response, "session", None):
                st.session_state.access_token = session_response.session.access_token
                st.session_state.refresh_token = session_response.session.refresh_token
                return

        session = supabase.auth.get_session()
        if session and getattr(session, "session", None):
            current_session = session.session
            if current_session and current_session.user:
                st.session_state.user = current_session.user
                st.session_state.access_token = current_session.access_token
                st.session_state.refresh_token = current_session.refresh_token
    except Exception:
        pass


# ------------------------------
# Database helpers
# ------------------------------
def upsert_profile(user):
    if not user:
        return
    payload = {"id": user.id, "email": user.email}
    return supabase.table("profiles").upsert(payload).execute()



def ensure_default_project(user_id: str):
    resp = (
        supabase.table("projects")
        .select("id,name")
        .eq("user_id", user_id)
        .eq("name", "General")
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    if not rows:
        supabase.table("projects").insert(
            {
                "user_id": user_id,
                "name": "General",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
        ).execute()



def create_project(user_id: str, name: str):
    return (
        supabase.table("projects")
        .insert(
            {
                "user_id": user_id,
                "name": name.strip(),
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
        )
        .execute()
    )



def get_projects(user_id: str):
    return (
        supabase.table("projects")
        .select("*")
        .eq("user_id", user_id)
        .order("updated_at", desc=True)
        .execute()
    )



def rename_project(project_id: str, user_id: str, new_name: str):
    return (
        supabase.table("projects")
        .update({"name": new_name.strip(), "updated_at": now_iso()})
        .eq("id", project_id)
        .eq("user_id", user_id)
        .execute()
    )



def get_project_by_id(project_id: str, user_id: str):
    resp = (
        supabase.table("projects")
        .select("*")
        .eq("id", project_id)
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    return rows[0] if rows else None



def delete_project(project_id: str, user_id: str):
    project = get_project_by_id(project_id, user_id)
    if not project:
        return False, "Project not found."

    if project["name"] == "General":
        return False, "You cannot delete the default General project."

    items_resp = (
        supabase.table("saved_items")
        .select("id,screenshot_path")
        .eq("project_id", project_id)
        .eq("user_id", user_id)
        .execute()
    )
    items = items_resp.data or []

    for item in items:
        screenshot_path = item.get("screenshot_path")
        if screenshot_path:
            try:
                supabase.storage.from_("screenshots").remove([screenshot_path])
            except Exception:
                pass

    supabase.table("saved_items").delete().eq("project_id", project_id).eq("user_id", user_id).execute()
    supabase.table("projects").delete().eq("id", project_id).eq("user_id", user_id).execute()

    return True, f"Project '{project['name']}' deleted."



def save_item(
    user_id: str,
    project_id: str,
    item_type: str,
    title: str,
    input_context: str,
    output_text: str,
    screenshot_path: str = None,
    source_filename: str = None,
):
    supabase.table("projects").update({"updated_at": now_iso()}).eq("id", project_id).eq("user_id", user_id).execute()

    payload = {
        "user_id": user_id,
        "project_id": project_id,
        "item_type": clean_text_for_storage(item_type).strip(),
        "title": clean_text_for_storage(title).strip(),
        "input_context": clean_text_for_storage(input_context).strip(),
        "output_text": clean_text_for_storage(output_text).strip(),
        "screenshot_path": screenshot_path,
        "source_filename": clean_text_for_storage(source_filename) if source_filename else None,
        "created_at": now_iso(),
    }
    return supabase.table("saved_items").insert(payload).execute()



def get_project_items(user_id: str, project_id: str):
    resp = (
        supabase.table("saved_items")
        .select("*")
        .eq("user_id", user_id)
        .eq("project_id", project_id)
        .order("created_at", desc=True)
        .execute()
    )
    return resp.data or []



def get_jira_integration(user_id: str):
    resp = (
        supabase.table("jira_integrations")
        .select("*")
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    return rows[0] if rows else None


# ------------------------------
# Storage helpers
# ------------------------------
def upload_screenshot_to_storage(user_id: str, project_id: str, uploaded_file):
    if uploaded_file is None:
        return None

    ext = os.path.splitext(uploaded_file.name)[1] or ".bin"
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{ext}"
    storage_path = f"{user_id}/{project_id}/{filename}"

    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)

    supabase.storage.from_("screenshots").upload(
        path=storage_path,
        file=file_bytes,
        file_options={"content-type": uploaded_file.type},
    )
    return storage_path


# ------------------------------
# Jira helpers
# ------------------------------
def get_default_jira_issue_type(output_type):
    mapping = {
        "Bug Report": "Bug",
        "Test Cases": "Task",
        "Test Scenarios": "Story",
        "Requirement to User Story": "Story",
        "Acceptance Criteria Generator": "Task",
        "User Story + Acceptance Criteria + Traceability": "Story",
        "Business Requirement Breakdown": "Task",
        "Business Process Flow": "Task",
        "Data Flow Diagram": "Task",
        "Technical Task Breakdown": "Task",
        "API / Backend Tasks": "Task",
        "Developer Checklist": "Task",
        "Smart Code Review": "Bug",
        "Technical Flow Diagram": "Task",
    }
    return mapping.get(output_type, "Task")



def get_default_jira_labels(output_type):
    mapping = {
        "Bug Report": ["bug"],
        "Test Cases": ["task"],
        "Test Scenarios": ["story"],
        "Requirement to User Story": ["story"],
        "Acceptance Criteria Generator": ["acceptance-criteria"],
        "User Story + Acceptance Criteria + Traceability": ["story", "acceptance-criteria", "traceability"],
        "Business Requirement Breakdown": ["business-requirement"],
        "Business Process Flow": ["process-flow"],
        "Data Flow Diagram": ["data-flow"],
        "Technical Task Breakdown": ["technical-task"],
        "API / Backend Tasks": ["api-task"],
        "Developer Checklist": ["developer-checklist"],
        "Smart Code Review": ["code-review"],
        "Technical Flow Diagram": ["technical-flow"],
    }
    return mapping.get(output_type, ["task"])



def build_jira_description_doc(description_text):
    lines = [line.strip() for line in description_text.splitlines() if line.strip()]

    content_blocks = []
    for line in lines:
        content_blocks.append(
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": line[:3000]}],
            }
        )

    if not content_blocks:
        content_blocks = [
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": "No description provided."}],
            }
        ]

    return {"type": "doc", "version": 1, "content": content_blocks}



def create_jira_issue(
    jira_base_url,
    jira_email,
    jira_api_token,
    jira_project_key,
    summary,
    description,
    issue_type="Task",
    labels=None,
):
    if not all([jira_base_url, jira_email, jira_api_token, jira_project_key]):
        return False, "Jira settings are missing. Please connect your Jira account first."

    url = f"{jira_base_url.rstrip('/')}/rest/api/3/issue"
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    auth = (jira_email, jira_api_token)

    payload = {
        "fields": {
            "project": {"key": jira_project_key},
            "summary": summary,
            "description": build_jira_description_doc(description),
            "issuetype": {"name": issue_type},
            "labels": labels or [],
        }
    }

    try:
        response = requests.post(url, json=payload, headers=headers, auth=auth, timeout=60)
    except Exception as e:
        return False, f"Connection failed: {e}"

    if response.status_code in [200, 201]:
        data = response.json()
        return True, data.get("key", "Created")

    return False, response.text


# ------------------------------
# OpenAI generators
# ------------------------------
def generate_bug_report_with_optional_image(title, context, uploaded_file):
    text_prompt = f"""
You are a senior QA engineer.

Generate a professional bug report for the following issue.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "bug_report": {{
    "Title": "",
    "Description": "",
    "Steps to Reproduce": "",
    "Expected Result": "",
    "Actual Result": "",
    "Severity": "",
    "Priority": "",
    "Environment": "",
    "Observed UI / Screenshot Notes": "",
    "Assumptions": ""
  }}
}}
"""

    if uploaded_file is not None:
        image_data_url = encode_uploaded_image(uploaded_file)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": text_prompt},
                        {"type": "image_url", "image_url": {"url": image_data_url}},
                    ],
                }
            ],
        )
    else:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": text_prompt}],
        )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    bug_report = parsed["bug_report"]

    pretty_text = "\n".join([f"{k}: {v}" for k, v in bug_report.items()])
    df = pd.DataFrame([bug_report])

    return pretty_text, df



def generate_test_cases(title, context):
    prompt = f"""
You are a senior QA engineer.

Generate detailed QA test cases for the following issue, feature, or requirement.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "test_cases": [
    {{
      "Test Case ID": "TC_001",
      "Category": "Functional",
      "Scenario": "",
      "Preconditions": "",
      "Steps": "",
      "Expected Result": "",
      "Priority": "",
      "Type": "Positive"
    }}
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    test_cases = parsed["test_cases"]

    pretty_lines = []
    for tc in test_cases:
        pretty_lines.append(
            f"""Test Case ID: {tc.get('Test Case ID', '')}
Category: {tc.get('Category', '')}
Scenario: {tc.get('Scenario', '')}
Preconditions: {tc.get('Preconditions', '')}
Steps: {tc.get('Steps', '')}
Expected Result: {tc.get('Expected Result', '')}
Priority: {tc.get('Priority', '')}
Type: {tc.get('Type', '')}
"""
        )

    pretty_text = "\n" + ("\n" + "-" * 80 + "\n").join(pretty_lines)
    df = pd.DataFrame(test_cases)

    return pretty_text, df



def generate_test_scenarios(title, context):
    prompt = f"""
You are a senior QA engineer.

Generate high-level test scenarios based on the following business requirement, feature, or issue.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "test_scenarios": [
    {{
      "Scenario ID": "TS_001",
      "Category": "Functional",
      "Scenario": "",
      "Description": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    scenarios = parsed["test_scenarios"]

    pretty_lines = []
    for sc in scenarios:
        pretty_lines.append(
            f"""Scenario ID: {sc.get('Scenario ID', '')}
Category: {sc.get('Category', '')}
Scenario: {sc.get('Scenario', '')}
Description: {sc.get('Description', '')}
Priority: {sc.get('Priority', '')}
Notes: {sc.get('Notes', '')}
"""
        )

    pretty_text = "\n" + ("\n" + "-" * 80 + "\n").join(pretty_lines)
    df = pd.DataFrame(scenarios)

    return pretty_text, df



def generate_ba_output(title, context, output_type):
    prompt_map = {
        "Requirement to User Story": f"""
You are a senior business analyst.

Convert the following requirement into structured user stories.

Title: {title}
Requirement Details: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "user_stories": [
    {{
      "User Story ID": "US_001",
      "As a": "",
      "I want": "",
      "So that": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
""",
        "Acceptance Criteria Generator": f"""
You are a senior business analyst.

Generate clear acceptance criteria for the following requirement.

Title: {title}
Requirement Details: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "acceptance_criteria": [
    {{
      "AC ID": "AC_001",
      "Criteria": "",
      "Type": "",
      "Priority": ""
    }}
  ]
}}
""",
        "Business Requirement Breakdown": f"""
You are a senior business analyst.

Break down the following requirement into structured business requirement points.

Title: {title}
Requirement Details: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "requirement_breakdown": [
    {{
      "Section": "",
      "Details": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
""",
    }

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt_map[output_type]}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)

    if output_type == "Requirement to User Story":
        records = parsed["user_stories"]
        pretty = "\n\n".join(
            [
                f"""User Story ID: {r.get('User Story ID', '')}
As a: {r.get('As a', '')}
I want: {r.get('I want', '')}
So that: {r.get('So that', '')}
Priority: {r.get('Priority', '')}
Notes: {r.get('Notes', '')}"""
                for r in records
            ]
        )
    elif output_type == "Acceptance Criteria Generator":
        records = parsed["acceptance_criteria"]
        pretty = "\n\n".join(
            [
                f"""AC ID: {r.get('AC ID', '')}
Criteria: {r.get('Criteria', '')}
Type: {r.get('Type', '')}
Priority: {r.get('Priority', '')}"""
                for r in records
            ]
        )
    else:
        records = parsed["requirement_breakdown"]
        pretty = "\n\n".join(
            [
                f"""Section: {r.get('Section', '')}
Details: {r.get('Details', '')}
Priority: {r.get('Priority', '')}
Notes: {r.get('Notes', '')}"""
                for r in records
            ]
        )

    df = pd.DataFrame(records)
    return pretty, df



def generate_story_ac_traceability_output(title, context, source_filename=""):
    prompt = f"""
You are a senior business analyst.

Convert the requirement into structured user stories with detailed acceptance criteria and rich traceability.

Important rules:
1. Generate as many user stories as reasonably supported by the requirement.
2. For each story, generate as many acceptance criteria as possible when supported by the source.
3. For each story, generate ALL possible traceability references from the requirement text, file content, sections, rows, headings, or excerpts.
4. Do not return only one traceability line.
5. Traceability must be returned as a LIST under "Traceability".
6. Return rich business detail when available from the requirement.
7. If exact page or line numbers are not available, leave them blank.
8. If section or sheet names are not available, infer the best available section / heading / area.
9. Mapping Type must be one of:
   - explicit
   - derived
   - assumed
10. Confidence must be one of:
   - high
   - medium
   - low

Title: {title}
Requirement Details:
{context}

Source File Name: {source_filename}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "stories": [
    {{
      "User Story ID": "US_001",
      "Story Title": "",
      "As a": "",
      "I want": "",
      "So that": "",
      "Priority": "",
      "Notes": "",
      "Business Rules / Details": "",
      "Acceptance Criteria": [
        "AC 1",
        "AC 2",
        "AC 3"
      ],
      "Acceptance Criteria Summary": "",
      "Traceability": [
        {{
          "Source File": "",
          "Source Section": "",
          "Sheet Name": "",
          "Page": "",
          "Line Start": "",
          "Line End": "",
          "Column / Field": "",
          "Source Excerpt": "",
          "Mapping Type": "explicit",
          "Confidence": "high",
          "Traceability Notes": ""
        }}
      ]
    }}
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    stories = parsed["stories"]

    flattened_rows = []
    pretty_blocks = []

    for s in stories:
        acceptance_criteria = s.get("Acceptance Criteria", [])
        if isinstance(acceptance_criteria, list):
            ac_text = "\n".join([f"- {item}" for item in acceptance_criteria])
            ac_count = len(acceptance_criteria)
        else:
            ac_text = str(acceptance_criteria)
            ac_count = 1 if str(acceptance_criteria).strip() else 0

        traceability_items = s.get("Traceability", [])
        if not isinstance(traceability_items, list):
            traceability_items = []

        trace_lines = []
        for idx, t in enumerate(traceability_items, start=1):
            trace_lines.append(
                f"""[{idx}]
Source File: {t.get('Source File', '')}
Source Section: {t.get('Source Section', '')}
Sheet Name: {t.get('Sheet Name', '')}
Page: {t.get('Page', '')}
Line Start: {t.get('Line Start', '')}
Line End: {t.get('Line End', '')}
Column / Field: {t.get('Column / Field', '')}
Source Excerpt: {t.get('Source Excerpt', '')}
Mapping Type: {t.get('Mapping Type', '')}
Confidence: {t.get('Confidence', '')}
Traceability Notes: {t.get('Traceability Notes', '')}"""
            )

        traceability_text = "\n\n".join(trace_lines) if trace_lines else "No detailed traceability found."
        traceability_count = len(traceability_items)

        pretty_blocks.append(
            f"""User Story ID: {s.get('User Story ID', '')}
Story Title: {s.get('Story Title', '')}
As a: {s.get('As a', '')}
I want: {s.get('I want', '')}
So that: {s.get('So that', '')}
Priority: {s.get('Priority', '')}
Notes: {s.get('Notes', '')}
Business Rules / Details: {s.get('Business Rules / Details', '')}

Acceptance Criteria ({ac_count}):
{ac_text}

Acceptance Criteria Summary: {s.get('Acceptance Criteria Summary', '')}

Traceability References ({traceability_count}):
{traceability_text}"""
        )

        flattened_rows.append({
            "User Story ID": s.get("User Story ID", ""),
            "Story Title": s.get("Story Title", ""),
            "As a": s.get("As a", ""),
            "I want": s.get("I want", ""),
            "So that": s.get("So that", ""),
            "Priority": s.get("Priority", ""),
            "Notes": s.get("Notes", ""),
            "Business Rules / Details": s.get("Business Rules / Details", ""),
            "Acceptance Criteria Count": ac_count,
            "Acceptance Criteria": ac_text,
            "Acceptance Criteria Summary": s.get("Acceptance Criteria Summary", ""),
            "Traceability Count": traceability_count,
            "Traceability Details": traceability_text,
        })

    pretty_text = "\n\n" + ("\n" + "-" * 100 + "\n").join(pretty_blocks)
    df = pd.DataFrame(flattened_rows)

    preferred_order = [
        "User Story ID",
        "Story Title",
        "As a",
        "I want",
        "So that",
        "Priority",
        "Acceptance Criteria Count",
        "Traceability Count",
        "Acceptance Criteria Summary",
        "Notes",
        "Business Rules / Details",
        "Acceptance Criteria",
        "Traceability Details",
    ]
    existing_cols = [c for c in preferred_order if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in existing_cols]
    df = df[existing_cols + remaining_cols]

    return pretty_text, df



def generate_dev_output(title, context, output_type):
    prompt_map = {
        "Technical Task Breakdown": f"""
You are a senior software engineer.

Break down the following requirement into technical development tasks.

Title: {title}
Technical Context: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "technical_tasks": [
    {{
      "Task ID": "DEV_001",
      "Task": "",
      "Component": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
""",
        "API / Backend Tasks": f"""
You are a senior backend engineer.

Generate backend and API development tasks for the following requirement.

Title: {title}
Technical Context: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "api_tasks": [
    {{
      "Task ID": "API_001",
      "Task": "",
      "Endpoint / Service": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
""",
        "Developer Checklist": f"""
You are a senior software engineer.

Generate a developer checklist for the following requirement.

Title: {title}
Technical Context: {context}

Return ONLY valid JSON.
Do not add markdown.

Use this exact JSON structure:
{{
  "developer_checklist": [
    {{
      "Checklist Item": "",
      "Category": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
""",
    }

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt_map[output_type]}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)

    if output_type == "Technical Task Breakdown":
        records = parsed["technical_tasks"]
        pretty = "\n\n".join(
            [
                f"""Task ID: {r.get('Task ID', '')}
Task: {r.get('Task', '')}
Component: {r.get('Component', '')}
Priority: {r.get('Priority', '')}
Notes: {r.get('Notes', '')}"""
                for r in records
            ]
        )
    elif output_type == "API / Backend Tasks":
        records = parsed["api_tasks"]
        pretty = "\n\n".join(
            [
                f"""Task ID: {r.get('Task ID', '')}
Task: {r.get('Task', '')}
Endpoint / Service: {r.get('Endpoint / Service', '')}
Priority: {r.get('Priority', '')}
Notes: {r.get('Notes', '')}"""
                for r in records
            ]
        )
    else:
        records = parsed["developer_checklist"]
        pretty = "\n\n".join(
            [
                f"""Checklist Item: {r.get('Checklist Item', '')}
Category: {r.get('Category', '')}
Priority: {r.get('Priority', '')}
Notes: {r.get('Notes', '')}"""
                for r in records
            ]
        )

    df = pd.DataFrame(records)
    return pretty, df



def generate_flow_diagram_output(title, context, diagram_type):
    prompt = f"""
You are a senior business systems analyst and solution architect.

Based on the requirement below, generate a {diagram_type}.

Title: {title}
Requirement Details: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation outside JSON.

Use this exact JSON structure:
{{
  "diagram_output": {{
    "diagram_type": "{diagram_type}",
    "mermaid_code": "flowchart TD\\n    A[Start] --> B[Process]",
    "steps": [
      {{
        "Step ID": "STEP_001",
        "From": "",
        "To": "",
        "Action": "",
        "Decision": "",
        "Notes": ""
      }}
    ]
  }}
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    diagram_output = parsed["diagram_output"]
    mermaid_code = diagram_output.get("mermaid_code", "")
    steps = diagram_output.get("steps", [])

    pretty_text = f"""Diagram Type: {diagram_output.get('diagram_type', diagram_type)}

Mermaid Code:
{mermaid_code}
"""
    df = pd.DataFrame(steps)
    return pretty_text, df, mermaid_code



def generate_requirements_from_flow(uploaded_flow):
    prompt = """
Analyze this flow diagram or process file and generate concise business requirements.

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{
  "requirements": {
    "Process Summary": "",
    "What Happens from Start to Finish": [
      "Step 1",
      "Step 2",
      "Step 3"
    ],
    "Important Decisions": [
      "Decision 1",
      "Decision 2"
    ],
    "Test Data Needed": [
      "Data Point 1",
      "Data Point 2"
    ]
  }
}
"""

    if uploaded_flow.type in ["image/png", "image/jpg", "image/jpeg"]:
        image_data_url = encode_uploaded_image(uploaded_flow)
        uploaded_flow.seek(0)

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": image_data_url}},
                    ],
                }
            ],
        )
        content = response.choices[0].message.content

    elif uploaded_flow.type == "application/pdf":
        file_bytes = uploaded_flow.read()
        uploaded_flow.seek(0)

        uploaded_pdf = client.files.create(
            file=(uploaded_flow.name, file_bytes, "application/pdf"),
            purpose="user_data",
        )

        response = client.responses.create(
            model="gpt-4.1",
            input=[
                {
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": prompt},
                        {"type": "input_file", "file_id": uploaded_pdf.id},
                    ],
                }
            ],
        )
        content = response.output_text

    else:
        uploaded_text = read_uploaded_text_file(uploaded_flow)
        if not uploaded_text.strip():
            raise ValueError("This file type is uploaded successfully, but readable text could not be extracted.")

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": f"{prompt}\n\nFile Content:\n{uploaded_text[:15000]}",
                }
            ],
        )
        content = response.choices[0].message.content

    parsed = parse_json_response(content)
    req = parsed["requirements"]

    steps_text = "\n".join(
        [f"{idx + 1}. {step}" for idx, step in enumerate(req.get("What Happens from Start to Finish", []))]
    )
    decisions_text = "\n".join([f"- {item}" for item in req.get("Important Decisions", [])])
    test_data_text = "\n".join([f"- {item}" for item in req.get("Test Data Needed", [])])

    pretty_text = f"""Process Summary:
{req.get('Process Summary', '')}

What Happens from Start to Finish:
{steps_text}

Important Decisions:
{decisions_text}

Test Data Needed:
{test_data_text}
"""

    df = pd.DataFrame(
        [
            {
                "Process Summary": req.get("Process Summary", ""),
                "What Happens from Start to Finish": steps_text,
                "Important Decisions": decisions_text,
                "Test Data Needed": test_data_text,
            }
        ]
    )

    return pretty_text, df


# ------------------------------
# Output helpers
# ------------------------------
def dataframe_to_pretty_text(df: pd.DataFrame, output_type: str, mermaid_code: str = "") -> str:
    working_df = df.copy()

    if "Select" in working_df.columns:
        working_df = working_df.drop(columns=["Select"])

    records = working_df.to_dict(orient="records")

    if output_type == "Bug Report":
        if not records:
            return ""
        return "\n".join([f"{k}: {v}" for k, v in records[0].items()])

    if output_type == "User Story + Acceptance Criteria + Traceability":
        blocks = []
        for row in records:
            block = f"""User Story ID: {row.get('User Story ID', '')}
Story Title: {row.get('Story Title', '')}
As a: {row.get('As a', '')}
I want: {row.get('I want', '')}
So that: {row.get('So that', '')}
Priority: {row.get('Priority', '')}
Notes: {row.get('Notes', '')}
Business Rules / Details: {row.get('Business Rules / Details', '')}

Acceptance Criteria Count: {row.get('Acceptance Criteria Count', '')}
Acceptance Criteria:
{row.get('Acceptance Criteria', '')}

Acceptance Criteria Summary: {row.get('Acceptance Criteria Summary', '')}

Traceability Count: {row.get('Traceability Count', '')}
Traceability Details:
{row.get('Traceability Details', '')}"""
            blocks.append(block)

        return "\n\n" + ("\n" + "-" * 100 + "\n").join(blocks)

    blocks = []
    for row in records:
        block = "\n".join([f"{k}: {v}" for k, v in row.items()])
        blocks.append(block)

    body = "\n\n" + ("\n" + "-" * 80 + "\n").join(blocks) if blocks else ""

    if mermaid_code:
        return f"Mermaid Code:\n{mermaid_code}\n\nStructured Steps:{body}"

    return body



def set_workspace_output(
    workspace_name: str,
    output_type: str,
    title: str,
    result_text: str,
    df,
    base_name: str,
    sheet_name: str,
    mermaid_code: str = "",
):
    state = get_workspace_state(workspace_name)
    editable_df = df.copy() if isinstance(df, pd.DataFrame) else None

    if isinstance(editable_df, pd.DataFrame) and output_type != "Bug Report" and "Select" not in editable_df.columns:
        editable_df.insert(0, "Select", False)

    state["generated_type"] = output_type
    state["generated_title"] = title
    state["generated_text"] = result_text
    state["generated_df"] = df
    state["generated_base_name"] = base_name
    state["generated_sheet_name"] = sheet_name
    state["generated_mermaid_code"] = mermaid_code
    state["original_output_text"] = result_text
    state["editable_output_text"] = result_text
    state["original_output_df"] = df.copy() if isinstance(df, pd.DataFrame) else None
    state["editable_output_df"] = editable_df
    state["smart_review_result"] = None



def set_smart_code_review_output(workspace_name: str, title: str, review_result: dict):
    state = get_workspace_state(workspace_name)
    state["generated_type"] = "Smart Code Review"
    state["generated_title"] = title
    state["generated_text"] = json.dumps(review_result, indent=2)
    state["generated_df"] = None
    state["generated_base_name"] = f"{safe_filename(title)}_smart_code_review"
    state["generated_sheet_name"] = "Smart_Code_Review"
    state["generated_mermaid_code"] = ""
    state["original_output_text"] = state["generated_text"]
    state["editable_output_text"] = state["generated_text"]
    state["original_output_df"] = None
    state["editable_output_df"] = None
    state["smart_review_result"] = review_result



def show_download_buttons(df, pretty_text, base_name, sheet_name):
    csv_data = convert_df_to_csv(df)
    excel_data = convert_df_to_excel(df, sheet_name=sheet_name)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.download_button(
            label="Download TXT",
            data=pretty_text,
            file_name=f"{base_name}.txt",
            mime="text/plain",
            use_container_width=True,
        )

    with col2:
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"{base_name}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col3:
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=f"{base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )



def build_row_summary(row_dict: dict, output_type: str, fallback_title: str) -> str:
    candidates = {
        "Bug Report": ["Title"],
        "Test Cases": ["Scenario", "Test Case ID", "Title"],
        "Test Scenarios": ["Scenario", "Scenario ID", "Title"],
        "Requirement to User Story": ["I want", "User Story ID", "Section"],
        "Acceptance Criteria Generator": ["Criteria", "AC ID"],
        "User Story + Acceptance Criteria + Traceability": ["Story Title", "I want", "User Story ID"],
        "Business Requirement Breakdown": ["Section", "Details"],
        "Business Process Flow": ["Action", "Step ID"],
        "Data Flow Diagram": ["Action", "Step ID"],
        "Technical Task Breakdown": ["Task", "Task ID"],
        "API / Backend Tasks": ["Task", "Task ID", "Endpoint / Service"],
        "Developer Checklist": ["Checklist Item", "Category"],
        "Smart Code Review": ["Issue", "Issue ID"],
        "Technical Flow Diagram": ["Action", "Step ID"],
    }

    for key in candidates.get(output_type, []):
        value = str(row_dict.get(key, "")).strip()
        if value:
            return value[:200]

    for value in row_dict.values():
        text = str(value).strip()
        if text:
            return text[:200]

    return fallback_title



def build_story_ac_traceability_description(row_dict: dict) -> str:
    return f"""User Story:
Story Title: {row_dict.get('Story Title', '')}
As a: {row_dict.get('As a', '')}
I want: {row_dict.get('I want', '')}
So that: {row_dict.get('So that', '')}

Priority:
{row_dict.get('Priority', '')}

Notes:
{row_dict.get('Notes', '')}

Business Rules / Details:
{row_dict.get('Business Rules / Details', '')}

Acceptance Criteria Count:
{row_dict.get('Acceptance Criteria Count', '')}

Acceptance Criteria:
{row_dict.get('Acceptance Criteria', '')}

Acceptance Criteria Summary:
{row_dict.get('Acceptance Criteria Summary', '')}

Traceability Count:
{row_dict.get('Traceability Count', '')}

Traceability Details:
{row_dict.get('Traceability Details', '')}
"""



def render_row_level_jira(output_type: str, generated_title: str, edited_df: pd.DataFrame):
    jira_config = get_jira_integration(st.session_state.user.id) if st.session_state.user else None
    st.markdown("### Create in Jira")

    issue_type_options = ["Bug", "Task", "Story"]
    default_issue_type = get_default_jira_issue_type(output_type)
    default_labels = get_default_jira_labels(output_type)
    default_index = issue_type_options.index(default_issue_type) if default_issue_type in issue_type_options else 0

    jira_issue_type = st.selectbox(
        "Issue Type",
        issue_type_options,
        index=default_index,
        key=f"jira_issue_type_{output_type}",
    )

    if not jira_config:
        st.info("Open Settings page to connect JIRA.")
        return

    st.success(
        f"Connected: {jira_config.get('jira_base_url')} / {jira_config.get('jira_project_key')}"
    )

    selectable_df = edited_df.copy()

    if "Select" in selectable_df.columns:
        selected_rows = selectable_df[selectable_df["Select"] == True].drop(columns=["Select"], errors="ignore")
    else:
        selected_rows = selectable_df

    if st.button("Create Jira for Selected Rows", use_container_width=True, key=f"create_jira_rows_{output_type}"):
        if selected_rows.empty:
            st.warning("Please select at least one row.")
            return

        created_keys = []
        failed_rows = []

        with st.spinner("Creating Jira issues..."):
            for _, row in selected_rows.iterrows():
                row_dict = row.to_dict()
                summary = build_row_summary(row_dict, output_type, generated_title)
                if output_type == "User Story + Acceptance Criteria + Traceability":
                    description = build_story_ac_traceability_description(row_dict)
                else:
                    description = "\n".join([f"{k}: {v}" for k, v in row_dict.items()])

                success, message = create_jira_issue(
                    jira_base_url=jira_config["jira_base_url"],
                    jira_email=jira_config["jira_email"],
                    jira_api_token=jira_config["jira_api_token"],
                    jira_project_key=jira_config["jira_project_key"],
                    summary=summary,
                    description=description,
                    issue_type=jira_issue_type,
                    labels=default_labels,
                )

                if success:
                    created_keys.append(message)
                else:
                    failed_rows.append(f"{summary}: {message}")

        if created_keys:
            st.success(f"Created {len(created_keys)} Jira issue(s): {', '.join(created_keys)}")
        if failed_rows:
            st.error("Some Jira issues failed:\n\n" + "\n".join(failed_rows))



def render_single_jira(output_type: str, generated_title: str, description_text: str):
    jira_config = get_jira_integration(st.session_state.user.id) if st.session_state.user else None
    st.markdown("### Create in Jira")

    issue_type_options = ["Bug", "Task", "Story"]
    default_issue_type = get_default_jira_issue_type(output_type)
    default_labels = get_default_jira_labels(output_type)
    default_index = issue_type_options.index(default_issue_type) if default_issue_type in issue_type_options else 0

    jira_issue_type = st.selectbox(
        "Issue Type",
        issue_type_options,
        index=default_index,
        key=f"jira_issue_type_{output_type}",
    )

    if not jira_config:
        st.info("Open Settings page to connect JIRA.")
        return

    st.success(
        f"Connected: {jira_config.get('jira_base_url')} / {jira_config.get('jira_project_key')}"
    )

    if st.button("Create in Jira", use_container_width=True, key=f"create_jira_{output_type}"):
        with st.spinner("Creating Jira issue..."):
            success, message = create_jira_issue(
                jira_base_url=jira_config["jira_base_url"],
                jira_email=jira_config["jira_email"],
                jira_api_token=jira_config["jira_api_token"],
                jira_project_key=jira_config["jira_project_key"],
                summary=generated_title,
                description=description_text,
                issue_type=jira_issue_type,
                labels=default_labels,
            )

        if success:
            issue_url = f"{jira_config['jira_base_url'].rstrip('/')}/browse/{message}"
            st.success(f"Jira issue created successfully: {message}")
            st.markdown(f"[Open Jira Issue]({issue_url})")
        else:
            st.error(f"Failed to create Jira issue: {message}")



def render_workspace_output(workspace_name: str):
    state = get_workspace_state(workspace_name)
    generated_type = state["generated_type"]

    if generated_type == "Smart Code Review" and state["smart_review_result"]:
        st.subheader("Generated Smart Code Review")
        render_code_review_results(state["smart_review_result"])

        json_text = json.dumps(state["smart_review_result"], indent=2)
        st.download_button(
            label="Download Smart Code Review JSON",
            data=json_text,
            file_name=f"{state['generated_base_name']}.json",
            mime="application/json",
            use_container_width=True,
            key=f"download_smart_code_review_json_{workspace_name}",
        )
        return

    if generated_type and state["editable_output_df"] is not None:
        generated_title = state["generated_title"]
        base_name = state["generated_base_name"]
        sheet_name = state["generated_sheet_name"]
        mermaid_code = state["generated_mermaid_code"]

        st.subheader(f"Generated {generated_type}")

        if mermaid_code:
            st.markdown("#### Diagram Preview")
            render_mermaid_diagram_with_exports(mermaid_code, workspace_name, generated_title)
            show_mermaid_download_buttons(mermaid_code, base_name, generated_title)

        edited_df = st.data_editor(
            state["editable_output_df"],
            use_container_width=True,
            num_rows="dynamic",
            key=f"editor_{workspace_name}_{generated_type}",
        )
        state["editable_output_df"] = edited_df

        edited_text = st.text_area(
            "Edit Output",
            value=state["editable_output_text"],
            height=320,
            key=f"editable_output_box_{workspace_name}",
        )

        c1, c2, c3 = st.columns(3)

        with c1:
            render_copy_button(edited_text, "Copy Output")

        with c2:
            if st.button("Update Output", use_container_width=True, key=f"update_output_btn_{workspace_name}"):
                state["generated_text"] = edited_text
                state["editable_output_text"] = edited_text
                st.success("Output updated successfully.")

        with c3:
            if st.button("Reset Output", use_container_width=True, key=f"reset_output_btn_{workspace_name}"):
                state["generated_text"] = state["original_output_text"]
                state["editable_output_text"] = state["original_output_text"]
                state["editable_output_df"] = state["original_output_df"].copy()
                st.rerun()

        final_text_from_grid = dataframe_to_pretty_text(
            state["editable_output_df"],
            generated_type,
            mermaid_code=mermaid_code,
        )

        show_download_buttons(
            state["editable_output_df"].drop(columns=["Select"], errors="ignore"),
            final_text_from_grid,
            base_name,
            sheet_name,
        )

        if generated_type == "Bug Report":
            render_single_jira(generated_type, generated_title, edited_text)
        else:
            render_row_level_jira(generated_type, generated_title, state["editable_output_df"])


# ------------------------------
# Sidebar
# ------------------------------
def render_sidebar_projects(user):
    st.sidebar.markdown("### Project")

    projects = get_projects(user.id).data or []

    if not projects:
        ensure_default_project(user.id)
        projects = get_projects(user.id).data or []

    if not st.session_state.selected_project_id and projects:
        st.session_state.selected_project_id = projects[0]["id"]

    selected_project = next(
        (p for p in projects if p["id"] == st.session_state.selected_project_id),
        projects[0] if projects else None,
    )

    if selected_project:
        project_names = [p["name"] for p in projects]
        current_index = next(
            (idx for idx, p in enumerate(projects) if p["id"] == selected_project["id"]),
            0,
        )

        selected_name = st.sidebar.selectbox(
            "Select Project",
            project_names,
            index=current_index,
            key="sidebar_selected_project_box",
        )

        selected_project = next((p for p in projects if p["name"] == selected_name), projects[0])
        st.session_state.selected_project_id = selected_project["id"]

    with st.sidebar.expander("Project Actions", expanded=False):
        new_project = st.text_input(
            "New project",
            placeholder="Example: Salesforce Regression",
            key="sidebar_new_project_name",
        )

        if st.button("Add Project", use_container_width=True, key="sidebar_add_project_btn"):
            project_name = new_project.strip()
            if not project_name:
                st.warning("Enter a project name.")
            else:
                if any(p["name"].lower() == project_name.lower() for p in projects):
                    st.warning("Project already exists.")
                else:
                    create_project(user.id, project_name)
                    created = get_projects(user.id).data or []
                    match = next((p for p in created if p["name"] == project_name), None)
                    if match:
                        st.session_state.selected_project_id = match["id"]
                    st.rerun()

        if selected_project:
            rename_value = st.text_input(
                "Rename selected project",
                value="" if selected_project["name"] == "General" else selected_project["name"],
                key="sidebar_rename_project_value",
            )

            c1, c2 = st.columns(2)

            with c1:
                if st.button("Rename", use_container_width=True, key="sidebar_rename_project_btn"):
                    new_name = rename_value.strip()
                    if not new_name:
                        st.warning("Enter a new project name.")
                    elif selected_project["name"] == "General":
                        st.warning("You cannot rename the default General project.")
                    elif any(p["name"].lower() == new_name.lower() and p["id"] != selected_project["id"] for p in projects):
                        st.warning("A project with that name already exists.")
                    else:
                        rename_project(selected_project["id"], user.id, new_name)
                        st.success(f"Project renamed to '{new_name}'.")
                        st.rerun()

            with c2:
                if st.button("Delete", use_container_width=True, key="sidebar_delete_project_btn"):
                    success, msg = delete_project(selected_project["id"], user.id)
                    if success:
                        st.session_state.selected_project_id = None
                        st.success(msg)
                        st.rerun()
                    else:
                        st.warning(msg)

    project_items = get_project_items(user.id, selected_project["id"]) if selected_project else []
    return projects, selected_project, project_items


# ------------------------------
# Auth screen
# ------------------------------
def render_auth_screen():
    st.markdown('<div class="hero-title">Kaldi One</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="hero-subtitle">AI-powered workspace for QA, BA, Development, and flow-based requirements.</div>',
        unsafe_allow_html=True,
    )

    tab1, tab2, tab3 = st.tabs(["Login", "Sign Up", "Forgot Password"])

    with tab1:
        with st.form("login_form"):
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            submit = st.form_submit_button("Login", use_container_width=True)

            if submit:
                if not email.strip() or not password.strip():
                    st.error("Please enter email and password.")
                else:
                    try:
                        response = sign_in_user(email.strip(), password)
                        handle_login_success(response)
                    except Exception as e:
                        st.error(f"Login failed: {auth_error_text(e)}")

    with tab2:
        with st.form("signup_form"):
            email = st.text_input("Email", key="signup_email")
            password = st.text_input("Password", type="password", key="signup_password")
            submit = st.form_submit_button("Create Account", use_container_width=True)

            if submit:
                if not email.strip() or not password.strip():
                    st.error("Please enter email and password.")
                else:
                    email = email.strip()
                    password = password.strip()

                    try:
                        existing_login = sign_in_user(email, password)
                        if getattr(existing_login, "session", None):
                            st.warning("This account is already created. Please use Sign In.")
                        else:
                            st.info("This account already exists. Please sign in, confirm your email, or use Forgot Password.")
                    except Exception:
                        try:
                            signup_response = sign_up_user(email, password)
                            if getattr(signup_response, "session", None):
                                st.success("Account created successfully.")
                            else:
                                st.success("Account created successfully. Please confirm your email, then sign in.")
                        except Exception as signup_exc:
                            st.error(f"Sign up failed: {auth_error_text(signup_exc)}")

    with tab3:
        with st.form("forgot_password_form"):
            email = st.text_input("Email", key="forgot_email")
            submit = st.form_submit_button("Send Reset Link", use_container_width=True)

            if submit:
                if not email.strip():
                    st.error("Please enter your email.")
                else:
                    try:
                        send_password_reset_email(email.strip())
                        st.success("Password reset email sent. Please check your inbox.")
                    except Exception as e:
                        st.error(f"Could not send reset email: {auth_error_text(e)}")


# ------------------------------
# Workspace render helpers
# ------------------------------
def render_workspace_header():
    st.markdown("""
    <div class="top-header-row">
        <div class="top-header-left">
            <div class="hero-title">Kaldi One</div>
            <div class="hero-subtitle">
                AI-powered workspace for QA, BA, Development, and flow-based requirements.
            </div>
        </div>
        <div class="top-header-right">
            <a class="top-link-btn pro" href="/?view=pricing" target="_blank">⭐ Upgrade to Pro</a>
            <a class="top-link-btn support" href="mailto:kaldiglobal1008@gmail.com">Support</a>
            <a class="top-link-btn logout" href="?do_logout=1" target="_self">Logout</a>
        </div>
    </div>
    """, unsafe_allow_html=True)



def render_workspace_buttons():
    c1, c2, c3, c4 = st.columns(4)
    active = st.session_state.active_workspace

    with c1:
        if st.button("QA Workspace", key="btn_qa_workspace", use_container_width=True, type="primary" if active == QA_WORKSPACE else "secondary"):
            st.session_state.active_workspace = QA_WORKSPACE
            st.rerun()

    with c2:
        if st.button("BA Workspace", key="btn_ba_workspace", use_container_width=True, type="primary" if active == BA_WORKSPACE else "secondary"):
            st.session_state.active_workspace = BA_WORKSPACE
            st.rerun()

    with c3:
        if st.button("Dev Workspace", key="btn_dev_workspace", use_container_width=True, type="primary" if active == DEV_WORKSPACE else "secondary"):
            st.session_state.active_workspace = DEV_WORKSPACE
            st.rerun()

    with c4:
        if st.button("Flow to Requirement", key="btn_flow_workspace", use_container_width=True, type="primary" if active == FLOW_WORKSPACE else "secondary"):
            st.session_state.active_workspace = FLOW_WORKSPACE
            st.rerun()



def render_qa_workspace(user, selected_project):
    workspace_name = QA_WORKSPACE
    state = get_workspace_state(workspace_name)

    st.markdown('<div class="clean-card">', unsafe_allow_html=True)
    st.subheader("QA Workspace")

    title = st.text_input(
        "Title",
        placeholder="Example: Login button not working on Safari mobile",
        key="qa_title_input",
    )

    context = st.text_area(
        "Requirement Details",
        height=160,
        placeholder="Paste bug details, requirement details, acceptance criteria, or observations here...",
        key="qa_context_input",
    )

    uploaded_file = st.file_uploader(
        "Upload Supporting File (Optional)",
        type=ALL_UPLOAD_FILE_TYPES,
        key="qa_file_upload",
    )
    maybe_reset_workspace_for_new_upload(workspace_name, uploaded_file)

    output_type = st.selectbox(
        "What do you want to generate?",
        ["Bug Report", "Test Cases", "Test Scenarios"],
        key="qa_output_type_select",
    )

    uploaded_text = read_uploaded_text_file(uploaded_file)
    full_context = context.strip()

    if uploaded_text.strip():
        if full_context:
            full_context = f"{full_context}\n\nUploaded File Content:\n{uploaded_text}"
        else:
            full_context = uploaded_text

    is_form_valid = bool(title.strip()) and bool(full_context.strip())

    if not is_form_valid:
        st.info("Enter Title and Requirement Details, or upload a supporting file.")

    if st.button("Generate", disabled=not is_form_valid, key="qa_generate_btn"):
        try:
            screenshot_path = (
                upload_screenshot_to_storage(user.id, selected_project["id"], uploaded_file)
                if uploaded_file
                else None
            )

            with st.spinner(f"Generating {output_type}..."):
                if output_type == "Bug Report":
                    image_file = (
                        uploaded_file
                        if uploaded_file and uploaded_file.type in ["image/png", "image/jpg", "image/jpeg"]
                        else None
                    )
                    result_text, df = generate_bug_report_with_optional_image(title, full_context, image_file)
                    sheet_name = "Bug_Report"
                    base_name = f"{safe_filename(title)}_bug_report"
                elif output_type == "Test Cases":
                    result_text, df = generate_test_cases(title, full_context)
                    sheet_name = "Test_Cases"
                    base_name = f"{safe_filename(title)}_test_cases"
                else:
                    result_text, df = generate_test_scenarios(title, full_context)
                    sheet_name = "Test_Scenarios"
                    base_name = f"{safe_filename(title)}_test_scenarios"

            set_workspace_output(workspace_name, output_type, title, result_text, df, base_name, sheet_name)

            save_item(
                user_id=user.id,
                project_id=selected_project["id"],
                item_type=output_type,
                title=title,
                input_context=full_context,
                output_text=result_text,
                screenshot_path=screenshot_path,
                source_filename=uploaded_file.name if uploaded_file else None,
            )

            st.success(f"{output_type} generated successfully.")

        except Exception as e:
            st.error(f"Failed to generate {output_type}: {e}")

    st.markdown("</div>", unsafe_allow_html=True)

    if state["generated_type"]:
        st.markdown('<div class="result-card">', unsafe_allow_html=True)
        render_workspace_output(workspace_name)
        if st.button("Clear Current Output", use_container_width=True, key="clear_qa_output"):
            reset_workspace_state(workspace_name)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)



def render_ba_workspace(user, selected_project):
    workspace_name = BA_WORKSPACE
    state = get_workspace_state(workspace_name)

    st.markdown('<div class="clean-card">', unsafe_allow_html=True)
    st.subheader("BA Workspace")

    title = st.text_input(
        "Requirement / Feature Title",
        placeholder="Example: Patient account phone number validation",
        key="ba_title_input",
    )

    context = st.text_area(
        "Business Requirement Details",
        height=180,
        placeholder="Paste the requirement, business rules, and expected behavior here...",
        key="ba_context_input",
    )

    uploaded_ba_file = st.file_uploader(
        "Upload Requirement File (Optional)",
        type=ALL_UPLOAD_FILE_TYPES,
        key="ba_file_upload",
    )
    maybe_reset_workspace_for_new_upload(workspace_name, uploaded_ba_file)

    output_type = st.selectbox(
        "Choose Output",
        [
            "Requirement to User Story",
            "Acceptance Criteria Generator",
            "User Story + Acceptance Criteria + Traceability",
            "Business Requirement Breakdown",
            "Business Process Flow",
            "Data Flow Diagram",
        ],
        key="ba_output_select",
    )

    uploaded_text = read_uploaded_text_file(uploaded_ba_file)
    trimmed_uploaded_text = trim_text_for_prompt(uploaded_text, 15000)
    full_context = clean_text_for_storage(context).strip()

    if trimmed_uploaded_text.strip():
        if full_context:
            full_context = f"{full_context}\n\nUploaded File Content:\n{trimmed_uploaded_text}"
        else:
            full_context = trimmed_uploaded_text

    if uploaded_text and len(uploaded_text) > 15000:
        st.warning("Uploaded file is large, so only the first portion was used for generation.")

    is_valid = bool(title.strip()) and bool(full_context.strip())

    if not is_valid:
        st.info("Enter title and requirement content, or upload a requirement file.")

    if st.button("Generate BA Output", disabled=not is_valid, key="ba_generate_btn"):
        try:
            with st.spinner(f"Generating {output_type}..."):
                if output_type in ["Business Process Flow", "Data Flow Diagram"]:
                    result_text, df, mermaid_code = generate_flow_diagram_output(title, full_context, output_type)
                elif output_type == "User Story + Acceptance Criteria + Traceability":
                    result_text, df = generate_story_ac_traceability_output(
                        title,
                        full_context,
                        uploaded_ba_file.name if uploaded_ba_file else "",
                    )
                    mermaid_code = ""
                else:
                    result_text, df = generate_ba_output(title, full_context, output_type)
                    mermaid_code = ""

            base_name = f"{safe_filename(title)}_{safe_filename(output_type)}"
            set_workspace_output(
                workspace_name,
                output_type,
                title,
                result_text,
                df,
                base_name,
                "BA_Output",
                mermaid_code=mermaid_code,
            )

            save_item(
                user_id=user.id,
                project_id=selected_project["id"],
                item_type=output_type,
                title=title,
                input_context=full_context,
                output_text=result_text,
                source_filename=uploaded_ba_file.name if uploaded_ba_file else None,
            )

            st.success(f"{output_type} generated successfully.")

        except Exception as e:
            st.error(f"Failed to generate BA output: {e}")

    st.markdown("</div>", unsafe_allow_html=True)

    if state["generated_type"]:
        st.markdown('<div class="result-card">', unsafe_allow_html=True)
        render_workspace_output(workspace_name)
        if st.button("Clear Current Output", use_container_width=True, key="clear_ba_output"):
            reset_workspace_state(workspace_name)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)



def render_dev_workspace(user, selected_project):
    workspace_name = DEV_WORKSPACE
    state = get_workspace_state(workspace_name)

    st.markdown('<div class="clean-card">', unsafe_allow_html=True)
    st.subheader("Dev Workspace")

    title = st.text_input(
        "User Story / Requirement Title",
        placeholder="Example: Add validation for duplicate phone number",
        key="dev_title_input",
    )

    context = st.text_area(
        "Technical Context / Code",
        height=220,
        placeholder="Paste the user story, business rules, API notes, or code here...",
        key="dev_context_input",
    )

    uploaded_dev_file = st.file_uploader(
        "Upload Dev File (Optional)",
        type=DEV_UPLOAD_FILE_TYPES,
        key="dev_file_upload",
    )
    maybe_reset_workspace_for_new_upload(workspace_name, uploaded_dev_file)

    language = st.selectbox(
        "Language",
        ["Auto / General", "Python", "JavaScript", "TypeScript", "Java", "C#", "SQL", "JSON", "Other"],
        key="dev_language_select",
    )

    output_type = st.selectbox(
        "Choose Output",
        [
            "Technical Task Breakdown",
            "API / Backend Tasks",
            "Developer Checklist",
            "Smart Code Review",
            "Technical Flow Diagram",
        ],
        key="dev_output_select",
    )

    uploaded_code_text = read_uploaded_text_file(uploaded_dev_file)
    full_context = context.strip()

    if uploaded_code_text.strip():
        if full_context:
            full_context = f"{full_context}\n\nUploaded File Content:\n{uploaded_code_text}"
        else:
            full_context = uploaded_code_text

    if language and language != "Auto / General":
        full_context = f"Language: {language}\n\n{full_context}"

    is_valid = bool(title.strip()) and bool(full_context.strip())

    if not is_valid:
        st.info("Enter title and technical context, or upload a dev file.")

    if st.button("Generate Dev Output", disabled=not is_valid, key="dev_generate_btn"):
        try:
            with st.spinner(f"Generating {output_type}..."):
                if output_type == "Smart Code Review":
                    review_result = run_smart_code_review(full_context)
                    result_text = json.dumps(review_result, indent=2)
                    set_smart_code_review_output(workspace_name, title, review_result)
                    save_item(
                        user_id=user.id,
                        project_id=selected_project["id"],
                        item_type=output_type,
                        title=title,
                        input_context=full_context,
                        output_text=result_text,
                        source_filename=uploaded_dev_file.name if uploaded_dev_file else None,
                    )
                elif output_type == "Technical Flow Diagram":
                    result_text, df, mermaid_code = generate_flow_diagram_output(title, full_context, output_type)
                    base_name = f"{safe_filename(title)}_{safe_filename(output_type)}"
                    set_workspace_output(
                        workspace_name,
                        output_type,
                        title,
                        result_text,
                        df,
                        base_name,
                        "Dev_Output",
                        mermaid_code=mermaid_code,
                    )
                    save_item(
                        user_id=user.id,
                        project_id=selected_project["id"],
                        item_type=output_type,
                        title=title,
                        input_context=full_context,
                        output_text=result_text,
                        source_filename=uploaded_dev_file.name if uploaded_dev_file else None,
                    )
                else:
                    result_text, df = generate_dev_output(title, full_context, output_type)
                    base_name = f"{safe_filename(title)}_{safe_filename(output_type)}"
                    set_workspace_output(
                        workspace_name,
                        output_type,
                        title,
                        result_text,
                        df,
                        base_name,
                        "Dev_Output",
                        mermaid_code="",
                    )
                    save_item(
                        user_id=user.id,
                        project_id=selected_project["id"],
                        item_type=output_type,
                        title=title,
                        input_context=full_context,
                        output_text=result_text,
                        source_filename=uploaded_dev_file.name if uploaded_dev_file else None,
                    )

            st.success(f"{output_type} generated successfully.")

        except Exception as e:
            st.error(f"Failed to generate Dev output: {e}")

    st.markdown("</div>", unsafe_allow_html=True)

    if state["generated_type"]:
        st.markdown('<div class="result-card">', unsafe_allow_html=True)
        render_workspace_output(workspace_name)
        if st.button("Clear Current Output", use_container_width=True, key="clear_dev_output"):
            reset_workspace_state(workspace_name)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)



def render_flow_workspace():
    workspace_name = FLOW_WORKSPACE
    state = get_workspace_state(workspace_name)

    st.markdown('<div class="clean-card">', unsafe_allow_html=True)
    st.subheader("Flow to Requirement")

    uploaded_flow = st.file_uploader(
        "Upload Flow File (Optional)",
        type=ALL_UPLOAD_FILE_TYPES,
        key="flow_file_upload",
    )
    maybe_reset_workspace_for_new_upload(workspace_name, uploaded_flow)

    if uploaded_flow is not None:
        if uploaded_flow.type in ["image/png", "image/jpg", "image/jpeg"]:
            st.image(uploaded_flow, caption="Uploaded Flow Diagram", use_container_width=True)
        elif uploaded_flow.type == "application/pdf":
            st.info(f"PDF uploaded: {uploaded_flow.name}")
        else:
            st.info(f"Uploaded file: {uploaded_flow.name}")

    flow_btn = st.button(
        "Generate Requirements",
        disabled=uploaded_flow is None,
        key="generate_flow_requirements",
    )

    if uploaded_flow is None:
        st.info("Upload a flow diagram to enable Generate Requirements.")

    if flow_btn and uploaded_flow is not None:
        try:
            flow_title = uploaded_flow.name.rsplit(".", 1)[0]

            with st.spinner("Generating requirements from flow..."):
                result_text, df_flow = generate_requirements_from_flow(uploaded_flow)

            base_name = f"{safe_filename(flow_title)}_requirements"
            set_workspace_output(
                workspace_name,
                "Flow Requirements",
                flow_title,
                result_text,
                df_flow,
                base_name,
                "Flow_Requirements",
                mermaid_code="",
            )

            st.success("Requirements generated successfully.")

        except Exception as e:
            st.error(f"Failed to generate requirements from flow: {e}")

    st.markdown("</div>", unsafe_allow_html=True)

    if state["generated_type"]:
        st.markdown('<div class="result-card">', unsafe_allow_html=True)
        render_workspace_output(workspace_name)
        if st.button("Clear Flow Output", use_container_width=True, key="clear_flow_output"):
            reset_workspace_state(workspace_name)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)



def render_main_app():
    user = st.session_state.user
    if not user:
        render_auth_screen()
        return

    _, selected_project, _ = render_sidebar_projects(user)

    if not selected_project:
        st.warning("Please create or select a project first.")
        return

    render_workspace_header()
    render_workspace_buttons()

    workspace = st.session_s
