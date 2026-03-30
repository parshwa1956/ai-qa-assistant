import os
import json
import base64
from io import BytesIO
from datetime import datetime, timezone

import pandas as pd
import requests
import streamlit as st
from openai import OpenAI
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------
# Load environment
# ------------------------------
load_dotenv()

# ------------------------------
# Page config
# ------------------------------
st.set_page_config(
    page_title="Kaldi QA",
    page_icon="🧪",
    layout="wide",
)

# ------------------------------
# Clients / secrets
# ------------------------------
def get_secret_or_env(name: str, default=None):
    try:
        return st.secrets[name]
    except Exception:
        return os.getenv(name, default)


OPENAI_API_KEY = get_secret_or_env("OPENAI_API_KEY")
SUPABASE_URL = get_secret_or_env("SUPABASE_URL")
SUPABASE_KEY = get_secret_or_env("SUPABASE_KEY")

JIRA_BASE_URL = get_secret_or_env("JIRA_BASE_URL")
JIRA_EMAIL = get_secret_or_env("JIRA_EMAIL")
JIRA_API_TOKEN = get_secret_or_env("JIRA_API_TOKEN")
JIRA_PROJECT_KEY = get_secret_or_env("JIRA_PROJECT_KEY")

if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY not found in Streamlit secrets or .env.")
    st.stop()

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("SUPABASE_URL or SUPABASE_KEY not found in Streamlit secrets or .env.")
    st.stop()

client = OpenAI(api_key=OPENAI_API_KEY)


@st.cache_resource
def get_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


supabase = get_supabase()

# ------------------------------
# Light layout tweak
# ------------------------------
st.markdown(
    """
<style>
.block-container {
    max-width: 1600px;
    padding-top: 2rem;
    padding-left: 2rem;
    padding-right: 2rem;
}

.small-muted {
    color: #6b7280;
    font-size: 0.85rem;
}

.sidebar-project-title {
    font-weight: 600;
    font-size: 0.96rem;
    margin-top: 0.25rem;
    margin-bottom: 0.15rem;
}

.sidebar-section-label {
    color: #6b7280;
    font-size: 0.82rem;
    margin-top: 0.5rem;
    margin-bottom: 0.35rem;
    text-transform: uppercase;
    letter-spacing: 0.03em;
}

.sidebar-item-card {
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 8px 10px;
    background: #ffffff;
    margin-bottom: 8px;
}

.sidebar-project-card {
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 10px 12px;
    background: #fafafa;
    margin-bottom: 10px;
}

.auth-wrap {
    max-width: 560px;
    margin: 0 auto;
    padding-top: 40px;
}

.auth-card {
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 24px;
    background: #ffffff;
    box-shadow: 0 1px 10px rgba(0,0,0,0.03);
}
</style>
""",
    unsafe_allow_html=True,
)

# ------------------------------
# Session state init
# ------------------------------
def init_session_state():
    defaults = {
        "user": None,
        "access_token": None,
        "selected_project_id": None,
        "generated_type": None,
        "generated_text": "",
        "generated_df": None,
        "generated_title": "",
        "generated_base_name": "",
        "generated_sheet_name": "",
        "flow_generated_text": "",
        "flow_generated_title": "",
        "flow_generated_df": None,
        "flow_generated_base_name": "",
        "flow_uploaded_name": "",
        "auth_checked": False,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


init_session_state()

# ------------------------------
# Utility helpers
# ------------------------------
def safe_filename(text):
    cleaned = "".join(c if c.isalnum() or c in (" ", "_", "-") else "" for c in text)
    cleaned = cleaned.strip().replace(" ", "_").lower()
    return cleaned if cleaned else "ai_output"


def encode_uploaded_image(uploaded_file):
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    encoded = base64.b64encode(file_bytes).decode("utf-8")
    mime_type = uploaded_file.type
    return f"data:{mime_type};base64,{encoded}"


def convert_df_to_csv(df):
    return df.to_csv(index=False).encode("utf-8")


def convert_df_to_excel(df, sheet_name="AI_Output"):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
        header_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_alignment

        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for cell in column_cells:
                cell.alignment = wrap_alignment
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 15), 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

        worksheet.freeze_panes = "A2"

    return output.getvalue()


def parse_json_response(content):
    content = content.strip()
    if content.startswith("```"):
        content = content.replace("```json", "").replace("```", "").strip()
    return json.loads(content)


def item_matches_search(item, search_term: str):
    if not search_term:
        return True

    search_term = search_term.lower().strip()
    haystack = " ".join(
        [
            str(item.get("title", "")),
            str(item.get("item_type", "")),
            str(item.get("output_text", ""))[:2000],
            str(item.get("source_filename", "")),
            str(item.get("input_context", ""))[:1000],
        ]
    ).lower()

    return search_term in haystack


def now_iso():
    return datetime.now(timezone.utc).isoformat()

# ------------------------------
# Auth helpers
# ------------------------------
def sign_up_user(email: str, password: str):
    return supabase.auth.sign_up({"email": email, "password": password})


def sign_in_user(email: str, password: str):
    return supabase.auth.sign_in_with_password({"email": email, "password": password})


PASSWORD_RESET_REDIRECT = get_secret_or_env("PASSWORD_RESET_REDIRECT", "http://localhost:8501")


def send_password_reset_email(email: str):
    return supabase.auth.reset_password_for_email(
        email,
        {"redirect_to": PASSWORD_RESET_REDIRECT},
    )


def update_logged_in_user_password(new_password: str):
    return supabase.auth.update_user({"password": new_password})


def auth_error_text(exc: Exception) -> str:
    text = str(exc)
    lower = text.lower()

    if "invalid login credentials" in lower:
        return "Unable to sign in. New user? Please use Sign Up. Already have an account? Use Forgot Password if needed."
    if "email not confirmed" in lower:
        return "Your account is created, but your email is not confirmed yet. Please check your inbox and confirm your email."
    if "user already registered" in lower:
        return "This account already exists. Please use Sign In or Forgot Password."
    if "invalid api key" in lower:
        return "Invalid Supabase configuration. Please check your Supabase URL and key in secrets.toml."
    return text


def sign_out_user():
    try:
        supabase.auth.sign_out()
    except Exception:
        pass

    for key in list(st.session_state.keys()):
        del st.session_state[key]
    init_session_state()


def load_user_from_existing_session():
    try:
        session = supabase.auth.get_session()
        if session and getattr(session, "session", None):
            current_session = session.session
            if current_session and current_session.user:
                st.session_state.user = current_session.user
                st.session_state.access_token = current_session.access_token
    except Exception:
        pass


def handle_login_success(auth_response):
    if getattr(auth_response, "user", None):
        st.session_state.user = auth_response.user
    if getattr(auth_response, "session", None):
        st.session_state.access_token = auth_response.session.access_token

    upsert_profile(st.session_state.user)
    ensure_default_project(st.session_state.user.id)
    st.rerun()

# ------------------------------
# Database helpers
# ------------------------------
def upsert_profile(user):
    if not user:
        return
    payload = {
        "id": user.id,
        "email": user.email,
    }
    return supabase.table("profiles").upsert(payload).execute()


def ensure_default_project(user_id: str):
    resp = (
        supabase.table("projects")
        .select("id,name")
        .eq("user_id", user_id)
        .eq("name", "General")
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    if not rows:
        supabase.table("projects").insert(
            {
                "user_id": user_id,
                "name": "General",
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
        ).execute()


def create_project(user_id: str, name: str):
    return (
        supabase.table("projects")
        .insert(
            {
                "user_id": user_id,
                "name": name.strip(),
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
        )
        .execute()
    )


def get_projects(user_id: str):
    return (
        supabase.table("projects")
        .select("*")
        .eq("user_id", user_id)
        .order("updated_at", desc=True)
        .execute()
    )


def rename_project(project_id: str, user_id: str, new_name: str):
    return (
        supabase.table("projects")
        .update(
            {
                "name": new_name.strip(),
                "updated_at": now_iso(),
            }
        )
        .eq("id", project_id)
        .eq("user_id", user_id)
        .execute()
    )


def get_project_by_id(project_id: str, user_id: str):
    resp = (
        supabase.table("projects")
        .select("*")
        .eq("id", project_id)
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    rows = resp.data or []
    return rows[0] if rows else None


def delete_project(project_id: str, user_id: str):
    project = get_project_by_id(project_id, user_id)
    if not project:
        return False, "Project not found."

    if project["name"] == "General":
        return False, "You cannot delete the default General project."

    items_resp = (
        supabase.table("saved_items")
        .select("id,screenshot_path")
        .eq("project_id", project_id)
        .eq("user_id", user_id)
        .execute()
    )
    items = items_resp.data or []

    for item in items:
        screenshot_path = item.get("screenshot_path")
        if screenshot_path:
            try:
                supabase.storage.from_("screenshots").remove([screenshot_path])
            except Exception:
                pass

    supabase.table("saved_items").delete().eq("project_id", project_id).eq("user_id", user_id).execute()
    supabase.table("projects").delete().eq("id", project_id).eq("user_id", user_id).execute()

    return True, f"Project '{project['name']}' deleted."


def save_item(
    user_id: str,
    project_id: str,
    item_type: str,
    title: str,
    input_context: str,
    output_text: str,
    screenshot_path: str = None,
    source_filename: str = None,
):
    supabase.table("projects").update({"updated_at": now_iso()}).eq("id", project_id).eq("user_id", user_id).execute()

    payload = {
        "user_id": user_id,
        "project_id": project_id,
        "item_type": item_type,
        "title": title.strip(),
        "input_context": input_context.strip(),
        "output_text": output_text.strip(),
        "screenshot_path": screenshot_path,
        "source_filename": source_filename,
        "created_at": now_iso(),
    }
    return supabase.table("saved_items").insert(payload).execute()


def get_project_items(user_id: str, project_id: str):
    resp = (
        supabase.table("saved_items")
        .select("*")
        .eq("user_id", user_id)
        .eq("project_id", project_id)
        .order("created_at", desc=True)
        .execute()
    )
    return resp.data or []


def get_recent_items(user_id: str, limit: int = 10):
    resp = (
        supabase.table("saved_items")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    return resp.data or []


def delete_project_item(user_id: str, item_id: str):
    item_resp = (
        supabase.table("saved_items")
        .select("id,screenshot_path")
        .eq("id", item_id)
        .eq("user_id", user_id)
        .limit(1)
        .execute()
    )
    rows = item_resp.data or []
    if not rows:
        return

    screenshot_path = rows[0].get("screenshot_path")
    if screenshot_path:
        try:
            supabase.storage.from_("screenshots").remove([screenshot_path])
        except Exception:
            pass

    supabase.table("saved_items").delete().eq("id", item_id).eq("user_id", user_id).execute()


def get_item_df(item):
    item_type = item.get("item_type", "")
    output_text = item.get("output_text", "")

    try:
        if item_type == "Bug Report":
            data = {}
            for line in output_text.splitlines():
                if ":" in line:
                    key, val = line.split(":", 1)
                    data[key.strip()] = val.strip()
            return pd.DataFrame([data]) if data else pd.DataFrame([{"Output": output_text}])

        return pd.DataFrame([{"Output": output_text}])
    except Exception:
        return pd.DataFrame([{"Output": output_text}])

# ------------------------------
# Storage helpers
# ------------------------------
def upload_screenshot_to_storage(user_id: str, project_id: str, uploaded_file):
    if uploaded_file is None:
        return None

    ext = os.path.splitext(uploaded_file.name)[1] or ".bin"
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{ext}"
    storage_path = f"{user_id}/{project_id}/{filename}"

    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    uploaded_file.seek(0)

    supabase.storage.from_("screenshots").upload(
        path=storage_path,
        file=file_bytes,
        file_options={"content-type": uploaded_file.type},
    )
    return storage_path


def get_signed_screenshot_url(storage_path: str, expires_in: int = 3600):
    if not storage_path:
        return None
    try:
        res = supabase.storage.from_("screenshots").create_signed_url(storage_path, expires_in)
        if isinstance(res, dict):
            return res.get("signedURL") or res.get("signedUrl")
        return getattr(res, "get", lambda *_: None)("signedURL")
    except Exception:
        return None

# ------------------------------
# Jira helpers
# ------------------------------
def get_default_jira_issue_type(output_type):
    mapping = {
        "Bug Report": "Bug",
        "Test Cases": "Task",
        "Test Scenarios": "Story",
    }
    return mapping.get(output_type, "Task")


def get_default_jira_labels(output_type):
    mapping = {
        "Bug Report": ["bug"],
        "Test Cases": ["task"],
        "Test Scenarios": ["story"],
    }
    return mapping.get(output_type, ["task"])


def build_jira_description_doc(description_text):
    lines = [line.strip() for line in description_text.splitlines() if line.strip()]

    content_blocks = []
    for line in lines:
        content_blocks.append(
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": line[:3000]}],
            }
        )

    if not content_blocks:
        content_blocks = [
            {
                "type": "paragraph",
                "content": [{"type": "text", "text": "No description provided."}],
            }
        ]

    return {
        "type": "doc",
        "version": 1,
        "content": content_blocks,
    }


def create_jira_issue(summary, description, issue_type="Task", labels=None):
    if not all([JIRA_BASE_URL, JIRA_EMAIL, JIRA_API_TOKEN, JIRA_PROJECT_KEY]):
        return False, "Jira settings are missing in Streamlit secrets or .env."

    url = f"{JIRA_BASE_URL}/rest/api/3/issue"
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    auth = (JIRA_EMAIL, JIRA_API_TOKEN)

    payload = {
        "fields": {
            "project": {"key": JIRA_PROJECT_KEY},
            "summary": summary,
            "description": build_jira_description_doc(description),
            "issuetype": {"name": issue_type},
            "labels": labels or [],
        }
    }

    response = requests.post(url, json=payload, headers=headers, auth=auth, timeout=60)

    if response.status_code in [200, 201]:
        data = response.json()
        return True, data.get("key", "Created")

    return False, response.text

# ------------------------------
# OpenAI generators
# ------------------------------
def generate_bug_report_with_optional_image(title, context, uploaded_file):
    text_prompt = f"""
You are a senior QA engineer.

Generate a professional bug report for the following issue.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "bug_report": {{
    "Title": "",
    "Description": "",
    "Steps to Reproduce": "",
    "Expected Result": "",
    "Actual Result": "",
    "Severity": "",
    "Priority": "",
    "Environment": "",
    "Observed UI / Screenshot Notes": "",
    "Assumptions": ""
  }}
}}
"""

    if uploaded_file is not None:
        image_data_url = encode_uploaded_image(uploaded_file)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": text_prompt},
                        {"type": "image_url", "image_url": {"url": image_data_url}},
                    ],
                }
            ],
        )
    else:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": text_prompt}],
        )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    bug_report = parsed["bug_report"]

    pretty_text = "\n".join([f"{k}: {v}" for k, v in bug_report.items()])
    df = pd.DataFrame([bug_report])

    return pretty_text, df


def generate_test_cases(title, context):
    prompt = f"""
You are a senior QA engineer.

Generate detailed QA test cases for the following issue, feature, or requirement.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "test_cases": [
    {{
      "Test Case ID": "TC_001",
      "Category": "Functional",
      "Scenario": "",
      "Preconditions": "",
      "Steps": "",
      "Expected Result": "",
      "Priority": "",
      "Type": "Positive"
    }}
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    test_cases = parsed["test_cases"]

    pretty_lines = []
    for tc in test_cases:
        pretty_lines.append(
            f"""Test Case ID: {tc.get('Test Case ID', '')}
Category: {tc.get('Category', '')}
Scenario: {tc.get('Scenario', '')}
Preconditions: {tc.get('Preconditions', '')}
Steps: {tc.get('Steps', '')}
Expected Result: {tc.get('Expected Result', '')}
Priority: {tc.get('Priority', '')}
Type: {tc.get('Type', '')}
"""
        )

    pretty_text = "\n" + ("\n" + "-" * 80 + "\n").join(pretty_lines)
    df = pd.DataFrame(test_cases)

    return pretty_text, df


def generate_test_scenarios(title, context):
    prompt = f"""
You are a senior QA engineer.

Generate high-level test scenarios based on the following business requirement, feature, or issue.

Title / Requirement / Feature: {title}
Context: {context}

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{{
  "test_scenarios": [
    {{
      "Scenario ID": "TS_001",
      "Category": "Functional",
      "Scenario": "",
      "Description": "",
      "Priority": "",
      "Notes": ""
    }}
  ]
}}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
    )

    content = response.choices[0].message.content
    parsed = parse_json_response(content)
    scenarios = parsed["test_scenarios"]

    pretty_lines = []
    for sc in scenarios:
        pretty_lines.append(
            f"""Scenario ID: {sc.get('Scenario ID', '')}
Category: {sc.get('Category', '')}
Scenario: {sc.get('Scenario', '')}
Description: {sc.get('Description', '')}
Priority: {sc.get('Priority', '')}
Notes: {sc.get('Notes', '')}
"""
        )

    pretty_text = "\n" + ("\n" + "-" * 80 + "\n").join(pretty_lines)
    df = pd.DataFrame(scenarios)

    return pretty_text, df


def reset_flow_output_if_new_file(uploaded_flow):
    if uploaded_flow is None:
        return

    if st.session_state.flow_uploaded_name != uploaded_flow.name:
        st.session_state.flow_generated_text = ""
        st.session_state.flow_generated_title = ""
        st.session_state.flow_generated_df = None
        st.session_state.flow_generated_base_name = ""
        st.session_state.flow_uploaded_name = uploaded_flow.name


def generate_requirements_from_flow(uploaded_flow):
    prompt = """
Analyze this flow diagram and generate concise business requirements.

Return ONLY valid JSON.
Do not add markdown.
Do not add explanation text.

Use this exact JSON structure:
{
  "requirements": {
    "Process Summary": "",
    "What Happens from Start to Finish": [
      "Step 1",
      "Step 2",
      "Step 3"
    ],
    "Important Decisions": [
      "Decision 1",
      "Decision 2"
    ],
    "Test Data Needed": [
      "Data Point 1",
      "Data Point 2"
    ]
  }
}
"""

    if uploaded_flow.type in ["image/png", "image/jpg", "image/jpeg"]:
        image_data_url = encode_uploaded_image(uploaded_flow)
        uploaded_flow.seek(0)

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": image_data_url}},
                    ],
                }
            ],
        )
        content = response.choices[0].message.content

    elif uploaded_flow.type == "application/pdf":
        file_bytes = uploaded_flow.read()
        uploaded_flow.seek(0)

        uploaded_pdf = client.files.create(
            file=(uploaded_flow.name, file_bytes, "application/pdf"),
            purpose="user_data",
        )

        response = client.responses.create(
            model="gpt-4.1",
            input=[
                {
                    "role": "user",
                    "content": [
                        {"type": "input_text", "text": prompt},
                        {"type": "input_file", "file_id": uploaded_pdf.id},
                    ],
                }
            ],
        )

        content = response.output_text
    else:
        raise ValueError("Unsupported file type. Please upload PNG, JPG, JPEG, or PDF.")

    parsed = parse_json_response(content)
    req = parsed["requirements"]

    steps_text = "\n".join([f"{idx + 1}. {step}" for idx, step in enumerate(req.get("What Happens from Start to Finish", []))])
    decisions_text = "\n".join([f"- {item}" for item in req.get("Important Decisions", [])])
    test_data_text = "\n".join([f"- {item}" for item in req.get("Test Data Needed", [])])

    pretty_text = f"""Process Summary:
{req.get('Process Summary', '')}

What Happens from Start to Finish:
{steps_text}

Important Decisions:
{decisions_text}

Test Data Needed:
{test_data_text}
"""

    df = pd.DataFrame(
        [
            {
                "Process Summary": req.get("Process Summary", ""),
                "What Happens from Start to Finish": steps_text,
                "Important Decisions": decisions_text,
                "Test Data Needed": test_data_text,
            }
        ]
    )

    return pretty_text, df

# ------------------------------
# Output / render helpers
# ------------------------------
def show_download_buttons(df, pretty_text, base_name, sheet_name):
    csv_data = convert_df_to_csv(df)
    excel_data = convert_df_to_excel(df, sheet_name=sheet_name)

    col1, col2, col3 = st.columns(3)

    with col1:
        st.download_button(
            label="Download TXT",
            data=pretty_text,
            file_name=f"{base_name}.txt",
            mime="text/plain",
            use_container_width=True,
        )

    with col2:
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"{base_name}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col3:
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=f"{base_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def set_current_output(output_type, title, result_text, df, base_name, sheet_name):
    st.session_state.generated_type = output_type
    st.session_state.generated_title = title
    st.session_state.generated_text = result_text
    st.session_state.generated_df = df
    st.session_state.generated_base_name = base_name
    st.session_state.generated_sheet_name = sheet_name


def load_project_item_into_current_output(item):
    output_type = item.get("item_type", "")
    title = item.get("title", "")
    text = item.get("output_text", "")

    if output_type == "Bug Report":
        sheet_name = "Bug_Report"
    elif output_type == "Test Cases":
        sheet_name = "Test_Cases"
    elif output_type == "Test Scenarios":
        sheet_name = "Test_Scenarios"
    else:
        sheet_name = "AI_Output"

    st.session_state.generated_type = output_type
    st.session_state.generated_title = title
    st.session_state.generated_text = text
    st.session_state.generated_df = get_item_df(item)
    st.session_state.generated_base_name = safe_filename(title)
    st.session_state.generated_sheet_name = sheet_name


def render_current_output():
    if st.session_state.generated_type and st.session_state.generated_df is not None:
        output_type = st.session_state.generated_type
        result_text = st.session_state.generated_text
        df = st.session_state.generated_df
        base_name = st.session_state.generated_base_name
        sheet_name = st.session_state.generated_sheet_name
        generated_title = st.session_state.generated_title

        st.subheader(f"Generated {output_type}")
        st.dataframe(df, use_container_width=True)

        st.text_area(
            "Copy Output",
            value=result_text,
            height=280,
            key=f"copy_output_{output_type}",
        )

        show_download_buttons(df, result_text, base_name, sheet_name)

        st.markdown("### Jira")

        default_issue_type = get_default_jira_issue_type(output_type)
        default_labels = get_default_jira_labels(output_type)

        issue_type_options = ["Bug", "Task", "Story"]
        default_index = issue_type_options.index(default_issue_type) if default_issue_type in issue_type_options else 0

        jira_issue_type = st.selectbox(
            "Issue Type",
            issue_type_options,
            index=default_index,
            key=f"jira_issue_type_{output_type}",
        )

        if st.button("Create in Jira", use_container_width=True, key=f"create_jira_{output_type}"):
            with st.spinner("Creating Jira issue..."):
                success, message = create_jira_issue(
                    summary=generated_title,
                    description=result_text,
                    issue_type=jira_issue_type,
                    labels=default_labels,
                )

            if success:
                issue_url = f"{JIRA_BASE_URL}/browse/{message}" if JIRA_BASE_URL else ""
                st.success(f"Jira issue created successfully: {message}")
                if issue_url:
                    st.markdown(f"[Open Jira Issue]({issue_url})")
            else:
                st.error(f"Failed to create Jira issue: {message}")


def render_flow_output():
    if st.session_state.flow_generated_df is not None:
        st.subheader("Generated Requirements")
        st.dataframe(st.session_state.flow_generated_df, use_container_width=True)

        st.text_area(
            "Copy Output",
            value=st.session_state.flow_generated_text,
            height=320,
            key="copy_output_flow_requirements",
        )

        show_download_buttons(
            st.session_state.flow_generated_df,
            st.session_state.flow_generated_text,
            st.session_state.flow_generated_base_name,
            "Flow_Requirements",
        )


def render_recent_history(user_id: str):
    items = get_recent_items(user_id, limit=10)
    if not items:
        return

    with st.expander("View Recent History"):
        for idx, item in enumerate(items, start=1):
            st.markdown(f"**{idx}. {item.get('item_type', '')} — {item.get('title', '')}**")

            df = get_item_df(item)
            if df is not None:
                st.dataframe(df, use_container_width=True)

            screenshot_path = item.get("screenshot_path")
            if screenshot_path:
                signed_url = get_signed_screenshot_url(screenshot_path)
                if signed_url:
                    st.image(
                        signed_url,
                        caption=f"Saved screenshot • {item.get('source_filename') or ''}",
                        use_container_width=True,
                    )

            st.text_area(
                f"History Output {idx}",
                value=item.get("output_text", ""),
                height=180,
                key=f"history_text_{idx}",
            )
            st.markdown("---")

# ------------------------------
# Sidebar
# ------------------------------
def render_sidebar_projects(user):
    st.sidebar.markdown("### Projects")
    st.sidebar.caption(user.email)

    if st.sidebar.button("Logout", use_container_width=True, key="sidebar_logout_btn"):
        sign_out_user()
        st.rerun()

    st.sidebar.markdown("---")

    new_project = st.sidebar.text_input(
        "New project",
        placeholder="Example: Salesforce Regression",
        key="sidebar_new_project_name",
    )

    if st.sidebar.button("Add Project", use_container_width=True, key="sidebar_add_project_btn"):
        project_name = new_project.strip()
        if not project_name:
            st.sidebar.warning("Enter a project name.")
        else:
            existing = get_projects(user.id).data or []
            if any(p["name"].lower() == project_name.lower() for p in existing):
                st.sidebar.warning("Project already exists.")
            else:
                create_project(user.id, project_name)
                created = get_projects(user.id).data or []
                match = next((p for p in created if p["name"] == project_name), None)
                if match:
                    st.session_state.selected_project_id = match["id"]
                st.rerun()

    projects = get_projects(user.id).data or []

    if not projects:
        st.sidebar.info("No projects found.")
        return [], None, []

    if not st.session_state.selected_project_id:
        st.session_state.selected_project_id = projects[0]["id"]

    selected_project = next((p for p in projects if p["id"] == st.session_state.selected_project_id), None)
    if not selected_project:
        selected_project = projects[0]
        st.session_state.selected_project_id = selected_project["id"]

    project_names = [p["name"] for p in projects]
    current_index = next((idx for idx, p in enumerate(projects) if p["id"] == selected_project["id"]), 0)

    selected_name = st.sidebar.selectbox(
        "Select Project",
        project_names,
        index=current_index,
        key="sidebar_selected_project_box",
    )

    selected_project = next((p for p in projects if p["name"] == selected_name), projects[0])
    st.session_state.selected_project_id = selected_project["id"]

    rename_value = st.sidebar.text_input(
        "Rename selected project",
        value="" if selected_project["name"] == "General" else selected_project["name"],
        key="sidebar_rename_project_value",
    )

    rename_col, delete_col = st.sidebar.columns(2)
    with rename_col:
        if st.button("Rename", use_container_width=True, key="sidebar_rename_project_btn"):
            new_name = rename_value.strip()
            if not new_name:
                st.sidebar.warning("Enter a new project name.")
            elif selected_project["name"] == "General":
                st.sidebar.warning("You cannot rename the default General project.")
            elif any(p["name"].lower() == new_name.lower() and p["id"] != selected_project["id"] for p in projects):
                st.sidebar.warning("A project with that name already exists.")
            else:
                rename_project(selected_project["id"], user.id, new_name)
                st.sidebar.success(f"Project renamed to '{new_name}'.")
                st.rerun()

    with delete_col:
        if st.button("Delete", use_container_width=True, key="sidebar_delete_project_btn"):
            success, msg = delete_project(selected_project["id"], user.id)
            if success:
                st.session_state.selected_project_id = None
                st.sidebar.success(msg)
                st.rerun()
            else:
                st.sidebar.warning(msg)

    project_items = get_project_items(user.id, selected_project["id"])

    st.sidebar.markdown(
        f"""
        <div class="sidebar-project-card">
            <div class="sidebar-project-title">{selected_project['name']}</div>
            <div class="small-muted">{len(project_items)} saved item(s)</div>
            <div class="small-muted">Updated: {selected_project.get('updated_at', '-') or '-'}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    search_term = st.sidebar.text_input(
        "Search saved items",
        placeholder="Search title, type, text...",
        key="sidebar_search_items",
    )

    st.sidebar.markdown('<div class="sidebar-section-label">Recent</div>', unsafe_allow_html=True)

    filtered_items = [item for item in project_items if item_matches_search(item, search_term)]

    if not filtered_items:
        st.sidebar.caption("No saved items found.")
        return projects, selected_project, project_items

    for item in filtered_items[:20]:
        st.sidebar.markdown(
            f"""
            <div class="sidebar-item-card">
                <strong>{item.get('title', '')}</strong><br>
                <span class="small-muted">{item.get('item_type', '')} • {item.get('created_at', '')}</span><br>
                <span class="small-muted">{item.get('source_filename') or ''}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        screenshot_path = item.get("screenshot_path")
        if screenshot_path:
            signed_url = get_signed_screenshot_url(screenshot_path)
            if signed_url:
                st.sidebar.image(signed_url, use_container_width=True)

        open_col, delete_col = st.sidebar.columns(2)
        with open_col:
            if st.button("Open", key=f"sidebar_open_item_{item['id']}", use_container_width=True):
                load_project_item_into_current_output(item)
                st.rerun()
        with delete_col:
            if st.button("Delete", key=f"sidebar_delete_item_{item['id']}", use_container_width=True):
                delete_project_item(user.id, item["id"])
                st.rerun()

    return projects, selected_project, project_items

# ------------------------------
# Auth screen
# ------------------------------
def render_auth_screen():
    st.markdown('<div class="auth-wrap">', unsafe_allow_html=True)
    st.markdown('<div class="auth-card">', unsafe_allow_html=True)

    st.title("Kaldi QA")
    st.caption("AI-powered test case generation and QA automation")
    st.write("Generate bug reports, test cases, high-level test scenarios, and flow-based requirements using AI.")

    if st.session_state.get("user"):
        with st.expander("Set new password"):
            with st.form("update_password_form"):
                new_password = st.text_input("New password", type="password", key="new_password")
                confirm_password = st.text_input("Confirm new password", type="password", key="confirm_new_password")
                pw_submit = st.form_submit_button("Update Password", use_container_width=True)

                if pw_submit:
                    if not new_password.strip():
                        st.error("Please enter a new password.")
                    elif new_password != confirm_password:
                        st.error("Passwords do not match.")
                    else:
                        try:
                            update_logged_in_user_password(new_password.strip())
                            st.success("Password updated successfully.")
                        except Exception as e:
                            st.error(f"Password update failed: {auth_error_text(e)}")

    tab1, tab2, tab3 = st.tabs(["Login", "Sign Up", "Forgot Password"])

    with tab1:
        with st.form("login_form"):
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            submit = st.form_submit_button("Login", use_container_width=True)
            st.caption("New here? Use Sign Up. Forgot your password? Use Forgot Password.")

            if submit:
                if not email.strip() or not password.strip():
                    st.error("Please enter email and password.")
                else:
                    try:
                        response = sign_in_user(email.strip(), password)
                        handle_login_success(response)
                    except Exception as e:
                        st.error(f"Login failed: {auth_error_text(e)}")

    with tab2:
        with st.form("signup_form"):
            email = st.text_input("Email", key="signup_email")
            password = st.text_input("Password", type="password", key="signup_password")
            submit = st.form_submit_button("Create Account", use_container_width=True)

            if submit:
                if not email.strip() or not password.strip():
                    st.error("Please enter email and password.")
                else:
                    email = email.strip()
                    password = password.strip()

                    try:
                        existing_login = sign_in_user(email, password)
                        if getattr(existing_login, "session", None):
                            st.warning("This account is already created. Please use Sign In.")
                        else:
                            st.info("This account already exists. Please sign in, confirm your email, or use Forgot Password.")
                    except Exception as login_exc:
                        login_text = str(login_exc).lower()

                        if "invalid login credentials" not in login_text and "email not confirmed" not in login_text:
                            st.error(f"Unable to validate account: {auth_error_text(login_exc)}")
                        else:
                            try:
                                signup_response = sign_up_user(email, password)

                                if getattr(signup_response, "session", None):
                                    st.success("Account created successfully.")
                                else:
                                    st.success("Account created successfully. Please confirm your email, then sign in.")
                            except Exception as signup_exc:
                                st.error(f"Sign up failed: {auth_error_text(signup_exc)}")

    with tab3:
        with st.form("forgot_password_form"):
            email = st.text_input("Email", key="forgot_email")
            submit = st.form_submit_button("Send Reset Link", use_container_width=True)

            if submit:
                if not email.strip():
                    st.error("Please enter your email.")
                else:
                    try:
                        send_password_reset_email(email.strip())
                        st.success("Password reset email sent. Please check your inbox.")
                    except Exception as e:
                        st.error(f"Could not send reset email: {auth_error_text(e)}")

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

# ------------------------------
# Main app UI
# ------------------------------
def render_main_app():
    user = st.session_state.user
    if not user:
        render_auth_screen()
        return

    projects, selected_project, project_items = render_sidebar_projects(user)

    st.title("Kaldi QA")
    st.caption("AI-powered test case generation and QA automation")
    st.write("Generate bug reports, test cases, high-level test scenarios, and flow-based requirements using AI.")

    tab1, tab2 = st.tabs(["QA Generator", "Flow to Requirements"])

    with tab1:
        title = st.text_input(
            "Title / Requirement / Feature *",
            placeholder="Example: Login button not working on Safari mobile",
        )

        context = st.text_area(
            "Context / Business Requirement Details *",
            height=150,
            placeholder="Paste bug details, requirement details, acceptance criteria, or observations here...",
        )

        uploaded_file = st.file_uploader(
            "Upload Screenshot (Optional)",
            type=["png", "jpg", "jpeg"],
            key="qa_screenshot_upload",
        )

        if uploaded_file is not None:
            st.image(uploaded_file, caption="Uploaded Screenshot", use_container_width=True)

        is_form_valid = bool(title.strip()) and bool(context.strip())

        if is_form_valid:
            st.success("Looks good. You can now generate outputs.")
        else:
            st.info("Enter Title and Context to enable Bug Report, Test Cases, and Test Scenarios.")

        col1, col2, col3 = st.columns(3)

        with col1:
            bug_btn = st.button("Generate Bug Report", disabled=not is_form_valid, use_container_width=True)

        with col2:
            case_btn = st.button("Generate Test Cases", disabled=not is_form_valid, use_container_width=True)

        with col3:
            scenario_btn = st.button("Generate Test Scenarios", disabled=not is_form_valid, use_container_width=True)

        if bug_btn:
            try:
                screenshot_path = upload_screenshot_to_storage(user.id, selected_project["id"], uploaded_file) if uploaded_file else None

                with st.spinner("Generating bug report..."):
                    result_text, df_bug = generate_bug_report_with_optional_image(title, context, uploaded_file)

                base_name = f"{safe_filename(title)}_bug_report"

                set_current_output("Bug Report", title, result_text, df_bug, base_name, "Bug_Report")

                save_item(
                    user_id=user.id,
                    project_id=selected_project["id"],
                    item_type="Bug Report",
                    title=title,
                    input_context=context,
                    output_text=result_text,
                    screenshot_path=screenshot_path,
                    source_filename=uploaded_file.name if uploaded_file else None,
                )

            except Exception as e:
                st.error(f"Failed to generate bug report: {e}")

        if case_btn:
            try:
                screenshot_path = upload_screenshot_to_storage(user.id, selected_project["id"], uploaded_file) if uploaded_file else None

                with st.spinner("Generating test cases..."):
                    result_text, df_cases = generate_test_cases(title, context)

                base_name = f"{safe_filename(title)}_test_cases"

                set_current_output("Test Cases", title, result_text, df_cases, base_name, "Test_Cases")

                save_item(
                    user_id=user.id,
                    project_id=selected_project["id"],
                    item_type="Test Cases",
                    title=title,
                    input_context=context,
                    output_text=result_text,
                    screenshot_path=screenshot_path,
                    source_filename=uploaded_file.name if uploaded_file else None,
                )

            except Exception as e:
                st.error(f"Failed to generate test cases: {e}")

        if scenario_btn:
            try:
                screenshot_path = upload_screenshot_to_storage(user.id, selected_project["id"], uploaded_file) if uploaded_file else None

                with st.spinner("Generating test scenarios..."):
                    result_text, df_scenarios = generate_test_scenarios(title, context)

                base_name = f"{safe_filename(title)}_test_scenarios"

                set_current_output("Test Scenarios", title, result_text, df_scenarios, base_name, "Test_Scenarios")

                save_item(
                    user_id=user.id,
                    project_id=selected_project["id"],
                    item_type="Test Scenarios",
                    title=title,
                    input_context=context,
                    output_text=result_text,
                    screenshot_path=screenshot_path,
                    source_filename=uploaded_file.name if uploaded_file else None,
                )

            except Exception as e:
                st.error(f"Failed to generate test scenarios: {e}")

        render_current_output()
        render_recent_history(user.id)

        if st.session_state.generated_type:
            if st.button("Clear Current Output", use_container_width=True, key="clear_qa_output"):
                st.session_state.generated_type = None
                st.session_state.generated_title = ""
                st.session_state.generated_text = ""
                st.session_state.generated_df = None
                st.session_state.generated_base_name = ""
                st.session_state.generated_sheet_name = ""
                st.rerun()

    with tab2:
        st.subheader("Flow Diagram to Requirements")
        st.caption("Upload a flow diagram and generate a simple explanation for general audience.")

        uploaded_flow = st.file_uploader(
            "Upload Flow Diagram",
            type=["png", "jpg", "jpeg", "pdf"],
            key="flow_diagram_upload",
        )

        if uploaded_flow is not None:
            reset_flow_output_if_new_file(uploaded_flow)

            if uploaded_flow.type in ["image/png", "image/jpg", "image/jpeg"]:
                st.image(uploaded_flow, caption="Uploaded Flow Diagram", use_container_width=True)
            elif uploaded_flow.type == "application/pdf":
                st.info(f"PDF uploaded: {uploaded_flow.name}")

        flow_btn = st.button(
            "Generate Requirements",
            disabled=uploaded_flow is None,
            use_container_width=True,
            key="generate_flow_requirements",
        )

        if uploaded_flow is None:
            st.info("Upload a flow diagram to enable Generate Requirements.")

        if flow_btn and uploaded_flow is not None:
            try:
                flow_title = uploaded_flow.name.rsplit(".", 1)[0]

                with st.spinner("Generating requirements from flow..."):
                    result_text, df_flow = generate_requirements_from_flow(uploaded_flow)

                st.session_state.flow_generated_title = flow_title
                st.session_state.flow_generated_text = result_text
                st.session_state.flow_generated_df = df_flow
                st.session_state.flow_generated_base_name = f"{safe_filename(flow_title)}_requirements"

            except Exception as e:
                st.error(f"Failed to generate requirements from flow: {e}")

        render_flow_output()

        if st.session_state.flow_generated_df is not None:
            if st.button("Clear Flow Output", use_container_width=True, key="clear_flow_output"):
                st.session_state.flow_generated_text = ""
                st.session_state.flow_generated_title = ""
                st.session_state.flow_generated_df = None
                st.session_state.flow_generated_base_name = ""
                st.rerun()

# ------------------------------
# Boot existing auth session once
# ------------------------------
if not st.session_state.auth_checked:
    load_user_from_existing_session()
    st.session_state.auth_checked = True

# ------------------------------
# Run app
# ------------------------------
render_main_app()
