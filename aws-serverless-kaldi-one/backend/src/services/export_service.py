import csv
import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_txt(output_text: str) -> bytes:
    return (output_text or "").encode("utf-8")


def export_csv(table_data: list[dict], output_text: str) -> bytes:
    buffer = io.StringIO()
    if table_data:
        writer = csv.DictWriter(buffer, fieldnames=list(table_data[0].keys()))
        writer.writeheader()
        writer.writerows(table_data)
    else:
        buffer.write(output_text or "")
    return buffer.getvalue().encode("utf-8")


def export_xlsx(table_data: list[dict], output_text: str, title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Export")[:31]

    if table_data:
        headers = list(table_data[0].keys())
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row_idx, row in enumerate(table_data, 2):
            for col_idx, key in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(row.get(key, "")))
    else:
        ws.cell(row=1, column=1, value=output_text or "")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def table_from_item(item: dict) -> list[dict]:
    output_json = item.get("outputJson")
    if isinstance(output_json, str):
        try:
            output_json = json.loads(output_json)
        except json.JSONDecodeError:
            output_json = None
    if not output_json:
        return []
    for key in (
        "test_cases",
        "test_scenarios",
        "user_stories",
        "acceptance_criteria",
        "requirement_breakdown",
        "technical_tasks",
        "api_tasks",
        "developer_checklist",
        "stories",
    ):
        if key in output_json and isinstance(output_json[key], list):
            return output_json[key]
    if "bug_report" in output_json:
        return [output_json["bug_report"]]
    if "diagram_output" in output_json:
        steps = output_json["diagram_output"].get("steps", [])
        return steps if isinstance(steps, list) else []
    return []
